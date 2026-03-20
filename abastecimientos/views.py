from datetime import date, datetime
from io import BytesIO
import re
import uuid
from statistics import median

from dateutil.relativedelta import relativedelta

from django import forms
from django.db.models import Sum
from django.db.models.deletion import ProtectedError
from django.db.models.functions import ExtractMonth, ExtractYear
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from .models import (
    AbastecimientoClaro,
    AbastecimientoColtrade,
    Canal,
    InventarioAbastecimiento,
    MetaAbastecimiento,
    ProductoAbastecimiento,
    PuntoVentaAbastecimiento,
    TransitoAbastecimiento,
    VentaAbastecimiento,
)
from user.jwt_utils import get_user_from_request


class CanalForm(forms.ModelForm):
    class Meta:
        model = Canal
        fields = ["canal_nombre"]
        widgets = {
            "canal_nombre": forms.TextInput(
                attrs={"placeholder": "Nombre del canal", "class": "form-control"}
            )
        }


class ProductoForm(forms.ModelForm):
    class Meta:
        model = ProductoAbastecimiento
        fields = ["id_producto", "nombre_producto", "marca", "id_canal"]
        widgets = {
            "id_producto": forms.TextInput(
                attrs={"placeholder": "ID producto", "class": "form-control"}
            ),
            "nombre_producto": forms.TextInput(
                attrs={"placeholder": "Nombre del producto", "class": "form-control"}
            ),
            "marca": forms.TextInput(
                attrs={"placeholder": "Marca", "class": "form-control"}
            ),
            "id_canal": forms.Select(attrs={"class": "form-control"}),
        }


class PuntoVentaForm(forms.ModelForm):
    class Meta:
        model = PuntoVentaAbastecimiento
        fields = ["id_puntoventa", "punto_venta", "canal_regional", "tipo", "id_canal"]
        widgets = {
            "id_puntoventa": forms.TextInput(
                attrs={"placeholder": "ID punto de venta", "class": "form-control"}
            ),
            "punto_venta": forms.TextInput(
                attrs={"placeholder": "Punto de venta", "class": "form-control"}
            ),
            "canal_regional": forms.TextInput(
                attrs={"placeholder": "Canal regional", "class": "form-control"}
            ),
            "tipo": forms.TextInput(
                attrs={"placeholder": "Tipo", "class": "form-control"}
            ),
            "id_canal": forms.Select(attrs={"class": "form-control"}),
        }


class InventarioForm(forms.ModelForm):
    class Meta:
        model = InventarioAbastecimiento
        fields = ["id_producto", "id_puntoventa", "cantidad_inventario", "id_canal"]
        widgets = {
            "id_producto": forms.Select(attrs={"class": "form-control"}),
            "id_puntoventa": forms.Select(attrs={"class": "form-control"}),
            "cantidad_inventario": forms.NumberInput(
                attrs={"placeholder": "Cantidad", "class": "form-control"}
            ),
            "id_canal": forms.Select(attrs={"class": "form-control"}),
        }


class MetaForm(forms.ModelForm):
    class Meta:
        model = MetaAbastecimiento
        fields = ["id_producto", "id_puntoventa", "cantidad_meta", "id_canal"]
        widgets = {
            "id_producto": forms.Select(attrs={"class": "form-control"}),
            "id_puntoventa": forms.Select(attrs={"class": "form-control"}),
            "cantidad_meta": forms.NumberInput(
                attrs={"placeholder": "Cantidad", "class": "form-control"}
            ),
            "id_canal": forms.Select(attrs={"class": "form-control"}),
        }


class TransitoForm(forms.ModelForm):
    class Meta:
        model = TransitoAbastecimiento
        fields = ["id_producto", "id_puntoventa", "cantidad_transito", "id_canal"]
        widgets = {
            "id_producto": forms.Select(attrs={"class": "form-control"}),
            "id_puntoventa": forms.Select(attrs={"class": "form-control"}),
            "cantidad_transito": forms.NumberInput(
                attrs={"placeholder": "Cantidad", "class": "form-control"}
            ),
            "id_canal": forms.Select(attrs={"class": "form-control"}),
        }


class VentaForm(forms.ModelForm):
    class Meta:
        model = VentaAbastecimiento
        fields = [
            "id_producto",
            "id_puntoventa",
            "cantidad_venta",
            "fecha_venta",
            "id_canal",
        ]
        widgets = {
            "id_producto": forms.Select(attrs={"class": "form-control"}),
            "id_puntoventa": forms.Select(attrs={"class": "form-control"}),
            "cantidad_venta": forms.NumberInput(
                attrs={"placeholder": "Cantidad", "class": "form-control"}
            ),
            "fecha_venta": forms.DateInput(
                attrs={"type": "date", "class": "form-control"}
            ),
            "id_canal": forms.Select(attrs={"class": "form-control"}),
        }


class AbastecimientoClaroForm(forms.ModelForm):
    class Meta:
        model = AbastecimientoClaro
        fields = [
            "material",
            "producto",
            "centro_costos",
            "nombre_punto",
            "inventario_claro",
            "transito_claro",
            "ventas_pasadas_claro",
            "ventas_actuales_claro",
            "sugerido_claro",
        ]
        widgets = {
            "material": forms.TextInput(attrs={"placeholder": "Material", "class": "form-control"}),
            "producto": forms.TextInput(attrs={"placeholder": "Producto", "class": "form-control"}),
            "centro_costos": forms.TextInput(
                attrs={"placeholder": "Centro costos", "class": "form-control"}
            ),
            "nombre_punto": forms.TextInput(
                attrs={"placeholder": "Nombre del punto", "class": "form-control"}
            ),
            "inventario_claro": forms.NumberInput(
                attrs={"placeholder": "Inventario", "class": "form-control"}
            ),
            "transito_claro": forms.NumberInput(
                attrs={"placeholder": "Transito", "class": "form-control"}
            ),
            "ventas_pasadas_claro": forms.NumberInput(
                attrs={"placeholder": "Ventas pasadas", "class": "form-control"}
            ),
            "ventas_actuales_claro": forms.NumberInput(
                attrs={"placeholder": "Ventas actuales", "class": "form-control"}
            ),
            "sugerido_claro": forms.NumberInput(
                attrs={"placeholder": "Sugerido", "class": "form-control"}
            ),
        }


class AbastecimientoColtradeForm(forms.ModelForm):
    class Meta:
        model = AbastecimientoColtrade
        fields = [
            "centro_costos",
            "punto_venta",
            "material",
            "producto",
            "marca",
            "ventas_actuales",
            "transitos",
            "inventario",
            "envio_inventario_3_meses",
            "sugerido_coltrade",
        ]
        widgets = {
            "centro_costos": forms.TextInput(
                attrs={"placeholder": "Centro costos", "class": "form-control"}
            ),
            "punto_venta": forms.TextInput(
                attrs={"placeholder": "Punto de venta", "class": "form-control"}
            ),
            "material": forms.TextInput(attrs={"placeholder": "Material", "class": "form-control"}),
            "producto": forms.TextInput(attrs={"placeholder": "Producto", "class": "form-control"}),
            "marca": forms.TextInput(attrs={"placeholder": "Marca", "class": "form-control"}),
            "ventas_actuales": forms.NumberInput(
                attrs={"placeholder": "Ventas actuales", "class": "form-control"}
            ),
            "transitos": forms.NumberInput(
                attrs={"placeholder": "Transitos", "class": "form-control"}
            ),
            "inventario": forms.NumberInput(
                attrs={"placeholder": "Inventario", "class": "form-control"}
            ),
            "envio_inventario_3_meses": forms.NumberInput(
                attrs={"placeholder": "Envio inv 3 meses", "class": "form-control"}
            ),
            "sugerido_coltrade": forms.NumberInput(
                attrs={"placeholder": "Sugerido", "class": "form-control"}
            ),
        }


def _build_xlsx_response(queryset, filename, sheet_name, exclude_fields=None):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    exclude_fields = set(exclude_fields or [])
    fields = [field for field in queryset.model._meta.fields if field.name not in exclude_fields]
    headers = [field.name for field in fields]
    ws.append(headers)
    for obj in queryset:
        ws.append([field.value_from_object(obj) for field in fields])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f"attachment; filename={filename}"
    return response


def _load_xlsx_rows(archivo):
    wb = load_workbook(archivo, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("El archivo está vacío.")
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    return headers, rows[1:]


def _row_is_empty(row):
    return not row or all(cell is None or str(cell).strip() == "" for cell in row)


def _get_cell(row, idx):
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None or str(value).strip() == "":
        return None
    try:
        return date.fromisoformat(str(value).strip())
    except ValueError as exc:
        raise ValueError(f"Fecha inválida: {value}") from exc

def canal_list(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    canales = Canal.objects.order_by("canal_nombre")
    form = CanalForm()
    return render(
        request,
        "canales_venta.html",
        {
            "canales": canales,
            "form": form,
            "edit_canal": None,
            "delete_error": None,
            "import_error": None,
        },
    )


def canal_create(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    if request.method != "POST":
        return redirect("abastecimientos:canal_list")
    form = CanalForm(request.POST)
    if form.is_valid():
        form.save()
        return redirect("abastecimientos:canal_list")
    canales = Canal.objects.order_by("canal_nombre")
    return render(
        request,
        "canales_venta.html",
        {
            "canales": canales,
            "form": form,
            "edit_canal": None,
            "delete_error": None,
            "import_error": None,
        },
    )


def canal_update(request, pk):
    if get_user_from_request(request) is None:
        return redirect("login")
    canal = get_object_or_404(Canal, pk=pk)
    if request.method == "POST":
        form = CanalForm(request.POST, instance=canal)
        if form.is_valid():
            form.save()
            return redirect("abastecimientos:canal_list")
    else:
        form = CanalForm(instance=canal)
    canales = Canal.objects.order_by("canal_nombre")
    return render(
        request,
        "canales_venta.html",
        {
            "canales": canales,
            "form": form,
            "edit_canal": canal,
            "delete_error": None,
            "import_error": None,
        },
    )


def canal_delete(request, pk):
    if get_user_from_request(request) is None:
        return redirect("login")
    canal = get_object_or_404(Canal, pk=pk)
    if request.method == "POST":
        try:
            canal.delete()
            return redirect("abastecimientos:canal_list")
        except ProtectedError:
            canales = Canal.objects.order_by("canal_nombre")
            form = CanalForm()
            return render(
                request,
                "canales_venta.html",
                {
                    "canales": canales,
                    "form": form,
                    "edit_canal": None,
                    "delete_error": "No se puede eliminar: tiene registros relacionados.",
                    "import_error": None,
                },
            )
    return redirect("abastecimientos:canal_list")


def canal_export(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    queryset = Canal.objects.order_by("id_canal")
    return _build_xlsx_response(queryset, "canales_abastecimiento.xlsx", "Canales")


def canal_import(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    if request.method != "POST":
        return redirect("abastecimientos:canal_list")
    archivo = request.FILES.get("archivo")
    if not archivo:
        return redirect("abastecimientos:canal_list")
    try:
        headers, rows = _load_xlsx_rows(archivo)
        if "canal_nombre" not in headers:
            raise ValueError("Falta la columna canal_nombre.")
        header_index = {name: idx for idx, name in enumerate(headers)}

        ids_in_file = []
        nombres_sin_id = []
        for row in rows:
            if _row_is_empty(row):
                continue
            canal_nombre = _get_cell(row, header_index.get("canal_nombre"))
            if canal_nombre is None or str(canal_nombre).strip() == "":
                continue
            canal_nombre = str(canal_nombre).strip()
            id_value = None
            if "id_canal" in header_index:
                id_value = _get_cell(row, header_index.get("id_canal"))
            if id_value is not None and str(id_value).strip() != "":
                ids_in_file.append(str(id_value).strip())
            else:
                nombres_sin_id.append(canal_nombre)

        existentes_por_id = {}
        if ids_in_file:
            existentes_por_id = {
                c.id_canal: c for c in Canal.objects.filter(id_canal__in=ids_in_file)
            }
        existentes_por_nombre = {}
        if nombres_sin_id:
            existentes_por_nombre = {
                c.canal_nombre: c
                for c in Canal.objects.filter(canal_nombre__in=nombres_sin_id)
            }

        canales_a_crear = {}
        canales_a_actualizar = {}
        nombres_a_crear = {}

        with transaction.atomic():
            for row in rows:
                if _row_is_empty(row):
                    continue
                canal_nombre = _get_cell(row, header_index.get("canal_nombre"))
                if canal_nombre is None or str(canal_nombre).strip() == "":
                    continue
                canal_nombre = str(canal_nombre).strip()
                id_value = None
                if "id_canal" in header_index:
                    id_value = _get_cell(row, header_index.get("id_canal"))
                if id_value is not None and str(id_value).strip() != "":
                    id_value = str(id_value).strip()
                else:
                    id_value = None

                if id_value:
                    if id_value in existentes_por_id:
                        canal = existentes_por_id[id_value]
                        canal.canal_nombre = canal_nombre
                        canales_a_actualizar[id_value] = canal
                    else:
                        canales_a_crear[id_value] = Canal(
                            id_canal=id_value, canal_nombre=canal_nombre
                        )
                else:
                    if canal_nombre in existentes_por_nombre:
                        continue
                    nombres_a_crear[canal_nombre] = Canal(canal_nombre=canal_nombre)

            if canales_a_crear or nombres_a_crear:
                Canal.objects.bulk_create(
                    list(canales_a_crear.values()) + list(nombres_a_crear.values()),
                    batch_size=500,
                )
            if canales_a_actualizar:
                Canal.objects.bulk_update(
                    list(canales_a_actualizar.values()),
                    ["canal_nombre"],
                    batch_size=500,
                )
        return redirect("abastecimientos:canal_list")
    except Exception as exc:
        canales = Canal.objects.order_by("canal_nombre")
        form = CanalForm()
        return render(
            request,
            "canales_venta.html",
            {
                "canales": canales,
                "form": form,
                "edit_canal": None,
                "delete_error": None,
                "import_error": f"Error al importar: {exc}",
            },
        )


def producto_list(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    productos = (
        ProductoAbastecimiento.objects.select_related("id_canal")
        .order_by("nombre_producto")
    )
    canales = Canal.objects.order_by("canal_nombre")
    form = ProductoForm()
    return render(
        request,
        "productos_venta.html",
        {
            "productos": productos,
            "canales": canales,
            "form": form,
            "edit_producto": None,
            "delete_error": None,
            "import_error": None,
        },
    )


def producto_create(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    if request.method != "POST":
        return redirect("abastecimientos:producto_list")
    form = ProductoForm(request.POST)
    if form.is_valid():
        form.save()
        return redirect("abastecimientos:producto_list")
    productos = (
        ProductoAbastecimiento.objects.select_related("id_canal")
        .order_by("nombre_producto")
    )
    canales = Canal.objects.order_by("canal_nombre")
    return render(
        request,
        "productos_venta.html",
        {
            "productos": productos,
            "canales": canales,
            "form": form,
            "edit_producto": None,
            "delete_error": None,
            "import_error": None,
        },
    )


def producto_update(request, pk):
    if get_user_from_request(request) is None:
        return redirect("login")
    producto = get_object_or_404(ProductoAbastecimiento, pk=pk)
    if request.method == "POST":
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            form.save()
            return redirect("abastecimientos:producto_list")
    else:
        form = ProductoForm(instance=producto)
    form.fields["id_producto"].disabled = True
    productos = (
        ProductoAbastecimiento.objects.select_related("id_canal")
        .order_by("nombre_producto")
    )
    canales = Canal.objects.order_by("canal_nombre")
    return render(
        request,
        "productos_venta.html",
        {
            "productos": productos,
            "canales": canales,
            "form": form,
            "edit_producto": producto,
            "delete_error": None,
            "import_error": None,
        },
    )


def producto_delete(request, pk):
    if get_user_from_request(request) is None:
        return redirect("login")
    producto = get_object_or_404(ProductoAbastecimiento, pk=pk)
    if request.method == "POST":
        try:
            producto.delete()
            return redirect("abastecimientos:producto_list")
        except ProtectedError:
            productos = (
                ProductoAbastecimiento.objects.select_related("id_canal")
                .order_by("nombre_producto")
            )
            canales = Canal.objects.order_by("canal_nombre")
            form = ProductoForm()
            return render(
                request,
                "productos_venta.html",
                {
                    "productos": productos,
                    "canales": canales,
                    "form": form,
                    "edit_producto": None,
                    "delete_error": "No se puede eliminar: tiene registros relacionados.",
                    "import_error": None,
                },
            )
    return redirect("abastecimientos:producto_list")


def producto_export(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    queryset = ProductoAbastecimiento.objects.order_by("id_producto")
    return _build_xlsx_response(
        queryset,
        "productos_abastecimiento.xlsx",
        "Productos",
        exclude_fields=["id_canal"],
    )


def producto_import(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    if request.method != "POST":
        return redirect("abastecimientos:producto_list")
    archivo = request.FILES.get("archivo")
    if not archivo:
        return redirect("abastecimientos:producto_list")
    try:
        headers, rows = _load_xlsx_rows(archivo)
        required = ["id_producto", "nombre_producto", "marca"]
        for name in required:
            if name not in headers:
                raise ValueError(f"Falta la columna {name}.")
        canal_id = request.POST.get("canal_id")
        if canal_id:
            canal_id = str(canal_id).strip()
        has_id_canal = "id_canal" in headers
        if not canal_id and not has_id_canal:
            raise ValueError("Seleccione un canal o incluya la columna id_canal.")
        header_index = {name: idx for idx, name in enumerate(headers)}

        canales_validos = set(Canal.objects.values_list("pk", flat=True))

        ids_existentes = [
            str(_get_cell(row, header_index["id_producto"])).strip()
            for row in rows
            if not _row_is_empty(row)
            and _get_cell(row, header_index["id_producto"]) not in [None, ""]
        ]

        productos_existentes = {
            p.id_producto: p
            for p in ProductoAbastecimiento.objects.filter(id_producto__in=ids_existentes)
        }

        productos_a_crear = {}
        productos_a_actualizar = {}

        with transaction.atomic():
            for row in rows:
                if _row_is_empty(row):
                    continue
                id_producto = _get_cell(row, header_index["id_producto"])
                if id_producto is None or str(id_producto).strip() == "":
                    continue
                nombre_producto = _get_cell(row, header_index["nombre_producto"])
                marca = _get_cell(row, header_index["marca"])
                id_canal_row = (
                    canal_id if canal_id else _get_cell(row, header_index.get("id_canal"))
                )

                id_producto = str(id_producto).strip()
                nombre_producto = "" if nombre_producto is None else str(nombre_producto).strip()
                marca = "" if marca is None else str(marca).strip()
                id_canal_row = "" if id_canal_row is None else str(id_canal_row).strip()

                if not id_canal_row:
                    raise ValueError("id_canal es requerido.")
                if id_canal_row not in canales_validos:
                    raise ValueError(f"Canal no existe: {id_canal_row}")

                if id_producto in productos_existentes:
                    prod = productos_existentes[id_producto]
                    prod.nombre_producto = nombre_producto
                    prod.marca = marca
                    prod.id_canal_id = id_canal_row
                    productos_a_actualizar[id_producto] = prod
                else:
                    productos_a_crear[id_producto] = ProductoAbastecimiento(
                        id_producto=id_producto,
                        nombre_producto=nombre_producto,
                        marca=marca,
                        id_canal_id=id_canal_row,
                    )

            if productos_a_crear:
                ProductoAbastecimiento.objects.bulk_create(
                    list(productos_a_crear.values()), batch_size=500
                )

            if productos_a_actualizar:
                ProductoAbastecimiento.objects.bulk_update(
                    list(productos_a_actualizar.values()),
                    ["nombre_producto", "marca", "id_canal"],
                    batch_size=500,
                )

        return redirect("abastecimientos:producto_list")
    except Exception as exc:
        productos = (
            ProductoAbastecimiento.objects.select_related("id_canal")
            .order_by("nombre_producto")
        )
        canales = Canal.objects.order_by("canal_nombre")
        form = ProductoForm()
        return render(
            request,
            "productos_venta.html",
            {
                "productos": productos,
                "canales": canales,
                "form": form,
                "edit_producto": None,
                "delete_error": None,
                "import_error": f"Error al importar: {exc}",
            },
        )


def producto_delete_by_canal(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    if request.method != "POST":
        return redirect("abastecimientos:producto_list")
    canal_id = request.POST.get("canal_id")
    if canal_id:
        canal_id = str(canal_id).strip()
    if canal_id:
        ProductoAbastecimiento.objects.filter(id_canal_id=canal_id).delete()
    return redirect("abastecimientos:producto_list")


def punto_venta_list(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    puntos_venta = (
        PuntoVentaAbastecimiento.objects.select_related("id_canal")
        .order_by("punto_venta")
    )
    canales = Canal.objects.order_by("canal_nombre")
    form = PuntoVentaForm()
    return render(
        request,
        "puntos_venta_abastecimiento.html",
        {
            "puntos_venta": puntos_venta,
            "canales": canales,
            "form": form,
            "edit_punto_venta": None,
            "delete_error": None,
            "import_error": None,
        },
    )


def punto_venta_create(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    if request.method != "POST":
        return redirect("abastecimientos:punto_venta_list")
    form = PuntoVentaForm(request.POST)
    if form.is_valid():
        form.save()
        return redirect("abastecimientos:punto_venta_list")
    puntos_venta = (
        PuntoVentaAbastecimiento.objects.select_related("id_canal")
        .order_by("punto_venta")
    )
    canales = Canal.objects.order_by("canal_nombre")
    return render(
        request,
        "puntos_venta_abastecimiento.html",
        {
            "puntos_venta": puntos_venta,
            "canales": canales,
            "form": form,
            "edit_punto_venta": None,
            "delete_error": None,
            "import_error": None,
        },
    )


def punto_venta_update(request, pk):
    if get_user_from_request(request) is None:
        return redirect("login")
    punto_venta = get_object_or_404(PuntoVentaAbastecimiento, pk=pk)
    if request.method == "POST":
        form = PuntoVentaForm(request.POST, instance=punto_venta)
        if form.is_valid():
            form.save()
            return redirect("abastecimientos:punto_venta_list")
    else:
        form = PuntoVentaForm(instance=punto_venta)
    form.fields["id_puntoventa"].disabled = True
    puntos_venta = (
        PuntoVentaAbastecimiento.objects.select_related("id_canal")
        .order_by("punto_venta")
    )
    canales = Canal.objects.order_by("canal_nombre")
    return render(
        request,
        "puntos_venta_abastecimiento.html",
        {
            "puntos_venta": puntos_venta,
            "canales": canales,
            "form": form,
            "edit_punto_venta": punto_venta,
            "delete_error": None,
            "import_error": None,
        },
    )


def punto_venta_delete(request, pk):
    if get_user_from_request(request) is None:
        return redirect("login")
    punto_venta = get_object_or_404(PuntoVentaAbastecimiento, pk=pk)
    if request.method == "POST":
        try:
            punto_venta.delete()
            return redirect("abastecimientos:punto_venta_list")
        except ProtectedError:
            puntos_venta = (
                PuntoVentaAbastecimiento.objects.select_related("id_canal")
                .order_by("punto_venta")
            )
            canales = Canal.objects.order_by("canal_nombre")
            form = PuntoVentaForm()
            return render(
                request,
                "puntos_venta_abastecimiento.html",
                {
                    "puntos_venta": puntos_venta,
                    "canales": canales,
                    "form": form,
                    "edit_punto_venta": None,
                    "delete_error": "No se puede eliminar: tiene registros relacionados.",
                    "import_error": None,
                },
            )
    return redirect("abastecimientos:punto_venta_list")


def punto_venta_export(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    queryset = PuntoVentaAbastecimiento.objects.order_by("id_puntoventa")
    return _build_xlsx_response(
        queryset,
        "puntos_venta_abastecimiento.xlsx",
        "Puntos de venta",
        exclude_fields=["id_canal"],
    )


def punto_venta_import(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    if request.method != "POST":
        return redirect("abastecimientos:punto_venta_list")
    archivo = request.FILES.get("archivo")
    if not archivo:
        return redirect("abastecimientos:punto_venta_list")
    try:
        headers, rows = _load_xlsx_rows(archivo)
        required = ["id_puntoventa", "punto_venta", "canal_regional", "tipo"]
        for name in required:
            if name not in headers:
                raise ValueError(f"Falta la columna {name}.")
        canal_id = request.POST.get("canal_id")
        if canal_id:
            canal_id = str(canal_id).strip()
        has_id_canal = "id_canal" in headers
        if not canal_id and not has_id_canal:
            raise ValueError("Seleccione un canal o incluya la columna id_canal.")
        header_index = {name: idx for idx, name in enumerate(headers)}

        canales_validos = set(Canal.objects.values_list("pk", flat=True))

        ids_existentes = [
            str(_get_cell(row, header_index["id_puntoventa"])).strip()
            for row in rows
            if not _row_is_empty(row)
            and _get_cell(row, header_index["id_puntoventa"]) not in [None, ""]
        ]

        puntos_existentes = {
            p.id_puntoventa: p
            for p in PuntoVentaAbastecimiento.objects.filter(id_puntoventa__in=ids_existentes)
        }

        puntos_a_crear = {}
        puntos_a_actualizar = {}

        with transaction.atomic():
            for row in rows:
                if _row_is_empty(row):
                    continue
                id_puntoventa = _get_cell(row, header_index["id_puntoventa"])
                if id_puntoventa is None or str(id_puntoventa).strip() == "":
                    continue
                punto_venta = _get_cell(row, header_index["punto_venta"])
                canal_regional = _get_cell(row, header_index["canal_regional"])
                tipo = _get_cell(row, header_index["tipo"])
                id_canal_row = (
                    canal_id if canal_id else _get_cell(row, header_index.get("id_canal"))
                )

                id_puntoventa = str(id_puntoventa).strip()
                punto_venta = "" if punto_venta is None else str(punto_venta).strip()
                canal_regional = (
                    "" if canal_regional is None else str(canal_regional).strip()
                )
                tipo = "" if tipo is None else str(tipo).strip()
                id_canal_row = "" if id_canal_row is None else str(id_canal_row).strip()

                if not id_canal_row:
                    raise ValueError("id_canal es requerido.")
                if id_canal_row not in canales_validos:
                    raise ValueError(f"Canal no existe: {id_canal_row}")

                if id_puntoventa in puntos_existentes:
                    punto = puntos_existentes[id_puntoventa]
                    punto.punto_venta = punto_venta
                    punto.canal_regional = canal_regional
                    punto.tipo = tipo
                    punto.id_canal_id = id_canal_row
                    puntos_a_actualizar[id_puntoventa] = punto
                else:
                    puntos_a_crear[id_puntoventa] = PuntoVentaAbastecimiento(
                        id_puntoventa=id_puntoventa,
                        punto_venta=punto_venta,
                        canal_regional=canal_regional,
                        tipo=tipo,
                        id_canal_id=id_canal_row,
                    )

            if puntos_a_crear:
                PuntoVentaAbastecimiento.objects.bulk_create(
                    list(puntos_a_crear.values()), batch_size=500
                )

            if puntos_a_actualizar:
                PuntoVentaAbastecimiento.objects.bulk_update(
                    list(puntos_a_actualizar.values()),
                    ["punto_venta", "canal_regional", "tipo", "id_canal"],
                    batch_size=500,
                )

        return redirect("abastecimientos:punto_venta_list")
    except Exception as exc:
        puntos_venta = (
            PuntoVentaAbastecimiento.objects.select_related("id_canal")
            .order_by("punto_venta")
        )
        canales = Canal.objects.order_by("canal_nombre")
        form = PuntoVentaForm()
        return render(
            request,
            "puntos_venta_abastecimiento.html",
            {
                "puntos_venta": puntos_venta,
                "canales": canales,
                "form": form,
                "edit_punto_venta": None,
                "delete_error": None,
                "import_error": f"Error al importar: {exc}",
            },
        )


def punto_venta_delete_by_canal(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    if request.method != "POST":
        return redirect("abastecimientos:punto_venta_list")
    canal_id = request.POST.get("canal_id")
    if canal_id:
        canal_id = str(canal_id).strip()
    if canal_id:
        PuntoVentaAbastecimiento.objects.filter(id_canal_id=canal_id).delete()
    return redirect("abastecimientos:punto_venta_list")


def inventario_list(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    inventarios = (
        InventarioAbastecimiento.objects.select_related(
            "id_producto", "id_puntoventa", "id_canal"
        )
        .order_by("id_inventario")
    )
    canales = Canal.objects.order_by("canal_nombre")
    form = InventarioForm()
    return render(
        request,
        "inventario_abastecimiento.html",
        {
            "inventarios": inventarios,
            "canales": canales,
            "form": form,
            "edit_inventario": None,
            "delete_error": None,
            "import_error": None,
        },
    )


def inventario_create(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    if request.method != "POST":
        return redirect("abastecimientos:inventario_list")
    form = InventarioForm(request.POST)
    if form.is_valid():
        form.save()
        return redirect("abastecimientos:inventario_list")
    inventarios = (
        InventarioAbastecimiento.objects.select_related(
            "id_producto", "id_puntoventa", "id_canal"
        )
        .order_by("id_inventario")
    )
    canales = Canal.objects.order_by("canal_nombre")
    return render(
        request,
        "inventario_abastecimiento.html",
        {
            "inventarios": inventarios,
            "canales": canales,
            "form": form,
            "edit_inventario": None,
            "delete_error": None,
            "import_error": None,
        },
    )


def inventario_update(request, pk):
    if get_user_from_request(request) is None:
        return redirect("login")
    inventario = get_object_or_404(InventarioAbastecimiento, pk=pk)
    if request.method == "POST":
        form = InventarioForm(request.POST, instance=inventario)
        if form.is_valid():
            form.save()
            return redirect("abastecimientos:inventario_list")
    else:
        form = InventarioForm(instance=inventario)
    inventarios = (
        InventarioAbastecimiento.objects.select_related(
            "id_producto", "id_puntoventa", "id_canal"
        )
        .order_by("id_inventario")
    )
    canales = Canal.objects.order_by("canal_nombre")
    return render(
        request,
        "inventario_abastecimiento.html",
        {
            "inventarios": inventarios,
            "canales": canales,
            "form": form,
            "edit_inventario": inventario,
            "delete_error": None,
            "import_error": None,
        },
    )


def inventario_delete(request, pk):
    if get_user_from_request(request) is None:
        return redirect("login")
    inventario = get_object_or_404(InventarioAbastecimiento, pk=pk)
    if request.method == "POST":
        try:
            inventario.delete()
            return redirect("abastecimientos:inventario_list")
        except ProtectedError:
            inventarios = (
                InventarioAbastecimiento.objects.select_related(
                    "id_producto", "id_puntoventa", "id_canal"
                )
                .order_by("id_inventario")
            )
            canales = Canal.objects.order_by("canal_nombre")
            form = InventarioForm()
            return render(
                request,
                "inventario_abastecimiento.html",
                {
                    "inventarios": inventarios,
                    "canales": canales,
                    "form": form,
                    "edit_inventario": None,
                    "delete_error": "No se puede eliminar: tiene registros relacionados.",
                    "import_error": None,
                },
            )
    return redirect("abastecimientos:inventario_list")


def inventario_export(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    queryset = InventarioAbastecimiento.objects.order_by("id_inventario")
    return _build_xlsx_response(
        queryset,
        "inventario_abastecimiento.xlsx",
        "Inventario",
        exclude_fields=["id_canal"],
    )

from django.db import transaction

def inventario_import(request):
    if get_user_from_request(request) is None:
        return redirect("login")

    if request.method != "POST":
        return redirect("abastecimientos:inventario_list")

    archivo = request.FILES.get("archivo")
    if not archivo:
        return redirect("abastecimientos:inventario_list")

    try:
        headers, rows = _load_xlsx_rows(archivo)

        required = ["id_producto", "id_puntoventa", "cantidad_inventario"]
        for name in required:
            if name not in headers:
                raise ValueError(f"Falta la columna {name}.")

        canal_id = request.POST.get("canal_id")
        if canal_id:
            canal_id = str(canal_id).strip()

        has_id_canal = "id_canal" in headers
        if not canal_id and not has_id_canal:
            raise ValueError("Seleccione un canal o incluya la columna id_canal.")

        header_index = {name: idx for idx, name in enumerate(headers)}

        # Traer IDs válidos una sola vez
        productos_validos = set(ProductoAbastecimiento.objects.values_list("pk", flat=True))
        puntoventas_validos = set(PuntoVentaAbastecimiento.objects.values_list("pk", flat=True))
        canales_validos = set(Canal.objects.values_list("pk", flat=True))

        inventarios_a_crear = []
        inventarios_a_actualizar = []

        ids_existentes = [
            int(_get_cell(row, header_index["id_inventario"]))
            for row in rows
            if "id_inventario" in header_index
            and _get_cell(row, header_index["id_inventario"]) not in [None, ""]
        ]

        inventarios_existentes = {
            inv.id_inventario: inv
            for inv in InventarioAbastecimiento.objects.filter(id_inventario__in=ids_existentes)
        }

        with transaction.atomic():
            for row in rows:
                if _row_is_empty(row):
                    continue

                id_inventario = (
                    _get_cell(row, header_index["id_inventario"])
                    if "id_inventario" in header_index
                    else None
                )

                id_producto = _get_cell(row, header_index["id_producto"])
                id_puntoventa = _get_cell(row, header_index["id_puntoventa"])
                cantidad = _get_cell(row, header_index["cantidad_inventario"])
                id_canal_row = canal_id if canal_id else _get_cell(row, header_index.get("id_canal"))

                id_producto = "" if id_producto is None else str(id_producto).strip()
                id_puntoventa = "" if id_puntoventa is None else str(id_puntoventa).strip()
                id_canal_row = "" if id_canal_row is None else str(id_canal_row).strip()

                if not id_producto or not id_puntoventa or not id_canal_row:
                    raise ValueError("Faltan campos requeridos en inventario.")

                if cantidad is None or str(cantidad).strip() == "":
                    raise ValueError("cantidad_inventario es requerida.")

                cantidad = int(cantidad)

                if id_producto not in productos_validos:
                    raise ValueError(f"Producto no existe: {id_producto}")

                if id_puntoventa not in puntoventas_validos:
                    raise ValueError(f"Punto de venta no existe: {id_puntoventa}")

                if id_canal_row not in canales_validos:
                    raise ValueError(f"Canal no existe: {id_canal_row}")

                if id_inventario is not None and str(id_inventario).strip() != "":
                    id_inventario = int(id_inventario)
                    if id_inventario in inventarios_existentes:
                        inv = inventarios_existentes[id_inventario]
                        inv.id_producto_id = id_producto
                        inv.id_puntoventa_id = id_puntoventa
                        inv.cantidad_inventario = cantidad
                        inv.id_canal_id = id_canal_row
                        inventarios_a_actualizar.append(inv)
                    else:
                        inventarios_a_crear.append(
                            InventarioAbastecimiento(
                                id_producto_id=id_producto,
                                id_puntoventa_id=id_puntoventa,
                                cantidad_inventario=cantidad,
                                id_canal_id=id_canal_row,
                            )
                        )
                else:
                    inventarios_a_crear.append(
                        InventarioAbastecimiento(
                            id_producto_id=id_producto,
                            id_puntoventa_id=id_puntoventa,
                            cantidad_inventario=cantidad,
                            id_canal_id=id_canal_row,
                        )
                    )

            if inventarios_a_crear:
                InventarioAbastecimiento.objects.bulk_create(inventarios_a_crear, batch_size=500)

            if inventarios_a_actualizar:
                InventarioAbastecimiento.objects.bulk_update(
                    inventarios_a_actualizar,
                    ["id_producto", "id_puntoventa", "cantidad_inventario", "id_canal"],
                    batch_size=500,
                )

        return redirect("abastecimientos:inventario_list")

    except Exception as exc:
        inventarios = (
            InventarioAbastecimiento.objects.select_related(
                "id_producto", "id_puntoventa", "id_canal"
            ).order_by("id_inventario")
        )
        canales = Canal.objects.order_by("canal_nombre")
        form = InventarioForm()
        return render(
            request,
            "inventario_abastecimiento.html",
            {
                "inventarios": inventarios,
                "canales": canales,
                "form": form,
                "edit_inventario": None,
                "delete_error": None,
                "import_error": f"Error al importar: {exc}",
            },
        )

def inventario_delete_by_canal(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    if request.method != "POST":
        return redirect("abastecimientos:inventario_list")
    canal_id = request.POST.get("canal_id")
    if canal_id:
        canal_id = str(canal_id).strip()
    if canal_id:
        InventarioAbastecimiento.objects.filter(id_canal_id=canal_id).delete()
    return redirect("abastecimientos:inventario_list")


def meta_list(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    metas = (
        MetaAbastecimiento.objects.select_related(
            "id_producto", "id_puntoventa", "id_canal"
        )
        .order_by("id_meta")
    )
    canales = Canal.objects.order_by("canal_nombre")
    form = MetaForm()
    return render(
        request,
        "meta_abastecimiento.html",
        {
            "metas": metas,
            "canales": canales,
            "form": form,
            "edit_meta": None,
            "delete_error": None,
            "import_error": None,
        },
    )


def meta_create(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    if request.method != "POST":
        return redirect("abastecimientos:meta_list")
    form = MetaForm(request.POST)
    if form.is_valid():
        form.save()
        return redirect("abastecimientos:meta_list")
    metas = (
        MetaAbastecimiento.objects.select_related(
            "id_producto", "id_puntoventa", "id_canal"
        )
        .order_by("id_meta")
    )
    canales = Canal.objects.order_by("canal_nombre")
    return render(
        request,
        "meta_abastecimiento.html",
        {
            "metas": metas,
            "canales": canales,
            "form": form,
            "edit_meta": None,
            "delete_error": None,
            "import_error": None,
        },
    )


def meta_update(request, pk):
    if get_user_from_request(request) is None:
        return redirect("login")
    meta = get_object_or_404(MetaAbastecimiento, pk=pk)
    if request.method == "POST":
        form = MetaForm(request.POST, instance=meta)
        if form.is_valid():
            form.save()
            return redirect("abastecimientos:meta_list")
    else:
        form = MetaForm(instance=meta)
    metas = (
        MetaAbastecimiento.objects.select_related(
            "id_producto", "id_puntoventa", "id_canal"
        )
        .order_by("id_meta")
    )
    canales = Canal.objects.order_by("canal_nombre")
    return render(
        request,
        "meta_abastecimiento.html",
        {
            "metas": metas,
            "canales": canales,
            "form": form,
            "edit_meta": meta,
            "delete_error": None,
            "import_error": None,
        },
    )


def meta_delete(request, pk):
    if get_user_from_request(request) is None:
        return redirect("login")
    meta = get_object_or_404(MetaAbastecimiento, pk=pk)
    if request.method == "POST":
        try:
            meta.delete()
            return redirect("abastecimientos:meta_list")
        except ProtectedError:
            metas = (
                MetaAbastecimiento.objects.select_related(
                    "id_producto", "id_puntoventa", "id_canal"
                )
                .order_by("id_meta")
            )
            canales = Canal.objects.order_by("canal_nombre")
            form = MetaForm()
            return render(
                request,
                "meta_abastecimiento.html",
                {
                    "metas": metas,
                    "canales": canales,
                    "form": form,
                    "edit_meta": None,
                    "delete_error": "No se puede eliminar: tiene registros relacionados.",
                    "import_error": None,
                },
            )
    return redirect("abastecimientos:meta_list")


def meta_export(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    queryset = MetaAbastecimiento.objects.order_by("id_meta")
    return _build_xlsx_response(
        queryset, "meta_abastecimiento.xlsx", "Metas", exclude_fields=["id_canal"]
    )


def meta_import(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    if request.method != "POST":
        return redirect("abastecimientos:meta_list")
    archivo = request.FILES.get("archivo")
    if not archivo:
        return redirect("abastecimientos:meta_list")
    try:
        headers, rows = _load_xlsx_rows(archivo)
        required = ["id_producto", "id_puntoventa", "cantidad_meta"]
        for name in required:
            if name not in headers:
                raise ValueError(f"Falta la columna {name}.")
        canal_id = request.POST.get("canal_id")
        if canal_id:
            canal_id = str(canal_id).strip()
        has_id_canal = "id_canal" in headers
        if not canal_id and not has_id_canal:
            raise ValueError("Seleccione un canal o incluya la columna id_canal.")
        header_index = {name: idx for idx, name in enumerate(headers)}

        productos_validos = set(ProductoAbastecimiento.objects.values_list("pk", flat=True))
        puntoventas_validos = set(PuntoVentaAbastecimiento.objects.values_list("pk", flat=True))
        canales_validos = set(Canal.objects.values_list("pk", flat=True))

        ids_existentes = [
            int(_get_cell(row, header_index["id_meta"]))
            for row in rows
            if "id_meta" in header_index
            and _get_cell(row, header_index["id_meta"]) not in [None, ""]
        ]

        metas_existentes = {
            m.id_meta: m
            for m in MetaAbastecimiento.objects.filter(id_meta__in=ids_existentes)
        }

        metas_a_crear = []
        metas_a_actualizar = []

        with transaction.atomic():
            for row in rows:
                if _row_is_empty(row):
                    continue

                id_meta = (
                    _get_cell(row, header_index["id_meta"])
                    if "id_meta" in header_index
                    else None
                )

                id_producto = _get_cell(row, header_index["id_producto"])
                id_puntoventa = _get_cell(row, header_index["id_puntoventa"])
                cantidad = _get_cell(row, header_index["cantidad_meta"])
                id_canal_row = (
                    canal_id if canal_id else _get_cell(row, header_index.get("id_canal"))
                )

                id_producto = "" if id_producto is None else str(id_producto).strip()
                id_puntoventa = "" if id_puntoventa is None else str(id_puntoventa).strip()
                id_canal_row = "" if id_canal_row is None else str(id_canal_row).strip()

                if not id_producto or not id_puntoventa or not id_canal_row:
                    raise ValueError("Faltan campos requeridos en metas.")
                if cantidad is None or str(cantidad).strip() == "":
                    raise ValueError("cantidad_meta es requerida.")

                cantidad = int(cantidad)

                if id_producto not in productos_validos:
                    raise ValueError(f"Producto no existe: {id_producto}")
                if id_puntoventa not in puntoventas_validos:
                    raise ValueError(f"Punto de venta no existe: {id_puntoventa}")
                if id_canal_row not in canales_validos:
                    raise ValueError(f"Canal no existe: {id_canal_row}")

                if id_meta is not None and str(id_meta).strip() != "":
                    id_meta = int(id_meta)
                    if id_meta in metas_existentes:
                        meta = metas_existentes[id_meta]
                        meta.id_producto_id = id_producto
                        meta.id_puntoventa_id = id_puntoventa
                        meta.cantidad_meta = cantidad
                        meta.id_canal_id = id_canal_row
                        metas_a_actualizar.append(meta)
                    else:
                        metas_a_crear.append(
                            MetaAbastecimiento(
                                id_producto_id=id_producto,
                                id_puntoventa_id=id_puntoventa,
                                cantidad_meta=cantidad,
                                id_canal_id=id_canal_row,
                            )
                        )
                else:
                    metas_a_crear.append(
                        MetaAbastecimiento(
                            id_producto_id=id_producto,
                            id_puntoventa_id=id_puntoventa,
                            cantidad_meta=cantidad,
                            id_canal_id=id_canal_row,
                        )
                    )

            if metas_a_crear:
                MetaAbastecimiento.objects.bulk_create(metas_a_crear, batch_size=500)

            if metas_a_actualizar:
                MetaAbastecimiento.objects.bulk_update(
                    metas_a_actualizar,
                    ["id_producto", "id_puntoventa", "cantidad_meta", "id_canal"],
                    batch_size=500,
                )

        return redirect("abastecimientos:meta_list")
    except Exception as exc:
        metas = (
            MetaAbastecimiento.objects.select_related(
                "id_producto", "id_puntoventa", "id_canal"
            )
            .order_by("id_meta")
        )
        canales = Canal.objects.order_by("canal_nombre")
        form = MetaForm()
        return render(
            request,
            "meta_abastecimiento.html",
            {
                "metas": metas,
                "canales": canales,
                "form": form,
                "edit_meta": None,
                "delete_error": None,
                "import_error": f"Error al importar: {exc}",
            },
        )


def meta_delete_by_canal(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    if request.method != "POST":
        return redirect("abastecimientos:meta_list")
    canal_id = request.POST.get("canal_id")
    if canal_id:
        canal_id = str(canal_id).strip()
    if canal_id:
        MetaAbastecimiento.objects.filter(id_canal_id=canal_id).delete()
    return redirect("abastecimientos:meta_list")


def transito_list(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    transitos = (
        TransitoAbastecimiento.objects.select_related(
            "id_producto", "id_puntoventa", "id_canal"
        )
        .order_by("id_transito")
    )
    canales = Canal.objects.order_by("canal_nombre")
    form = TransitoForm()
    return render(
        request,
        "transito_abastecimiento.html",
        {
            "transitos": transitos,
            "canales": canales,
            "form": form,
            "edit_transito": None,
            "delete_error": None,
            "import_error": None,
        },
    )


def transito_create(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    if request.method != "POST":
        return redirect("abastecimientos:transito_list")
    form = TransitoForm(request.POST)
    if form.is_valid():
        form.save()
        return redirect("abastecimientos:transito_list")
    transitos = (
        TransitoAbastecimiento.objects.select_related(
            "id_producto", "id_puntoventa", "id_canal"
        )
        .order_by("id_transito")
    )
    canales = Canal.objects.order_by("canal_nombre")
    return render(
        request,
        "transito_abastecimiento.html",
        {
            "transitos": transitos,
            "canales": canales,
            "form": form,
            "edit_transito": None,
            "delete_error": None,
            "import_error": None,
        },
    )


def transito_update(request, pk):
    if get_user_from_request(request) is None:
        return redirect("login")
    transito = get_object_or_404(TransitoAbastecimiento, pk=pk)
    if request.method == "POST":
        form = TransitoForm(request.POST, instance=transito)
        if form.is_valid():
            form.save()
            return redirect("abastecimientos:transito_list")
    else:
        form = TransitoForm(instance=transito)
    transitos = (
        TransitoAbastecimiento.objects.select_related(
            "id_producto", "id_puntoventa", "id_canal"
        )
        .order_by("id_transito")
    )
    canales = Canal.objects.order_by("canal_nombre")
    return render(
        request,
        "transito_abastecimiento.html",
        {
            "transitos": transitos,
            "canales": canales,
            "form": form,
            "edit_transito": transito,
            "delete_error": None,
            "import_error": None,
        },
    )


def transito_delete(request, pk):
    if get_user_from_request(request) is None:
        return redirect("login")
    transito = get_object_or_404(TransitoAbastecimiento, pk=pk)
    if request.method == "POST":
        try:
            transito.delete()
            return redirect("abastecimientos:transito_list")
        except ProtectedError:
            transitos = (
                TransitoAbastecimiento.objects.select_related(
                    "id_producto", "id_puntoventa", "id_canal"
                )
                .order_by("id_transito")
            )
            canales = Canal.objects.order_by("canal_nombre")
            form = TransitoForm()
            return render(
                request,
                "transito_abastecimiento.html",
                {
                    "transitos": transitos,
                    "canales": canales,
                    "form": form,
                    "edit_transito": None,
                    "delete_error": "No se puede eliminar: tiene registros relacionados.",
                    "import_error": None,
                },
            )
    return redirect("abastecimientos:transito_list")


def transito_export(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    queryset = TransitoAbastecimiento.objects.order_by("id_transito")
    return _build_xlsx_response(
        queryset,
        "transito_abastecimiento.xlsx",
        "Tránsitos",
        exclude_fields=["id_canal"],
    )


def transito_import(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    if request.method != "POST":
        return redirect("abastecimientos:transito_list")
    archivo = request.FILES.get("archivo")
    if not archivo:
        return redirect("abastecimientos:transito_list")
    try:
        headers, rows = _load_xlsx_rows(archivo)
        required = ["id_producto", "id_puntoventa", "cantidad_transito"]
        for name in required:
            if name not in headers:
                raise ValueError(f"Falta la columna {name}.")
        canal_id = request.POST.get("canal_id")
        if canal_id:
            canal_id = str(canal_id).strip()
        has_id_canal = "id_canal" in headers
        if not canal_id and not has_id_canal:
            raise ValueError("Seleccione un canal o incluya la columna id_canal.")
        header_index = {name: idx for idx, name in enumerate(headers)}

        productos_validos = set(ProductoAbastecimiento.objects.values_list("pk", flat=True))
        puntoventas_validos = set(PuntoVentaAbastecimiento.objects.values_list("pk", flat=True))
        canales_validos = set(Canal.objects.values_list("pk", flat=True))

        ids_existentes = [
            int(_get_cell(row, header_index["id_transito"]))
            for row in rows
            if "id_transito" in header_index
            and _get_cell(row, header_index["id_transito"]) not in [None, ""]
        ]

        transitos_existentes = {
            t.id_transito: t
            for t in TransitoAbastecimiento.objects.filter(id_transito__in=ids_existentes)
        }

        transitos_a_crear = []
        transitos_a_actualizar = []

        with transaction.atomic():
            for row in rows:
                if _row_is_empty(row):
                    continue

                id_transito = (
                    _get_cell(row, header_index["id_transito"])
                    if "id_transito" in header_index
                    else None
                )

                id_producto = _get_cell(row, header_index["id_producto"])
                id_puntoventa = _get_cell(row, header_index["id_puntoventa"])
                cantidad = _get_cell(row, header_index["cantidad_transito"])
                id_canal_row = (
                    canal_id if canal_id else _get_cell(row, header_index.get("id_canal"))
                )

                id_producto = "" if id_producto is None else str(id_producto).strip()
                id_puntoventa = "" if id_puntoventa is None else str(id_puntoventa).strip()
                id_canal_row = "" if id_canal_row is None else str(id_canal_row).strip()

                if not id_producto or not id_puntoventa or not id_canal_row:
                    raise ValueError("Faltan campos requeridos en transito.")
                if cantidad is None or str(cantidad).strip() == "":
                    raise ValueError("cantidad_transito es requerida.")

                cantidad = int(cantidad)

                if id_producto not in productos_validos:
                    raise ValueError(f"Producto no existe: {id_producto}")
                if id_puntoventa not in puntoventas_validos:
                    raise ValueError(f"Punto de venta no existe: {id_puntoventa}")
                if id_canal_row not in canales_validos:
                    raise ValueError(f"Canal no existe: {id_canal_row}")

                if id_transito is not None and str(id_transito).strip() != "":
                    id_transito = int(id_transito)
                    if id_transito in transitos_existentes:
                        transito = transitos_existentes[id_transito]
                        transito.id_producto_id = id_producto
                        transito.id_puntoventa_id = id_puntoventa
                        transito.cantidad_transito = cantidad
                        transito.id_canal_id = id_canal_row
                        transitos_a_actualizar.append(transito)
                    else:
                        transitos_a_crear.append(
                            TransitoAbastecimiento(
                                id_producto_id=id_producto,
                                id_puntoventa_id=id_puntoventa,
                                cantidad_transito=cantidad,
                                id_canal_id=id_canal_row,
                            )
                        )
                else:
                    transitos_a_crear.append(
                        TransitoAbastecimiento(
                            id_producto_id=id_producto,
                            id_puntoventa_id=id_puntoventa,
                            cantidad_transito=cantidad,
                            id_canal_id=id_canal_row,
                        )
                    )

            if transitos_a_crear:
                TransitoAbastecimiento.objects.bulk_create(transitos_a_crear, batch_size=500)

            if transitos_a_actualizar:
                TransitoAbastecimiento.objects.bulk_update(
                    transitos_a_actualizar,
                    ["id_producto", "id_puntoventa", "cantidad_transito", "id_canal"],
                    batch_size=500,
                )

        return redirect("abastecimientos:transito_list")
    except Exception as exc:
        transitos = (
            TransitoAbastecimiento.objects.select_related(
                "id_producto", "id_puntoventa", "id_canal"
            )
            .order_by("id_transito")
        )
        canales = Canal.objects.order_by("canal_nombre")
        form = TransitoForm()
        return render(
            request,
            "transito_abastecimiento.html",
            {
                "transitos": transitos,
                "canales": canales,
                "form": form,
                "edit_transito": None,
                "delete_error": None,
                "import_error": f"Error al importar: {exc}",
            },
        )


def transito_delete_by_canal(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    if request.method != "POST":
        return redirect("abastecimientos:transito_list")
    canal_id = request.POST.get("canal_id")
    if canal_id:
        canal_id = str(canal_id).strip()
    if canal_id:
        TransitoAbastecimiento.objects.filter(id_canal_id=canal_id).delete()
    return redirect("abastecimientos:transito_list")


def venta_list(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    ventas = (
        VentaAbastecimiento.objects.select_related(
            "id_producto", "id_puntoventa", "id_canal"
        )
        .order_by("-fecha_venta", "id_venta")
    )
    canales = Canal.objects.order_by("canal_nombre")
    form = VentaForm()
    return render(
        request,
        "venta_abastecimiento.html",
        {
            "ventas": ventas,
            "canales": canales,
            "form": form,
            "edit_venta": None,
            "delete_error": None,
            "import_error": None,
        },
    )


def venta_create(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    if request.method != "POST":
        return redirect("abastecimientos:venta_list")
    form = VentaForm(request.POST)
    if form.is_valid():
        form.save()
        return redirect("abastecimientos:venta_list")
    ventas = (
        VentaAbastecimiento.objects.select_related(
            "id_producto", "id_puntoventa", "id_canal"
        )
        .order_by("-fecha_venta", "id_venta")
    )
    canales = Canal.objects.order_by("canal_nombre")
    return render(
        request,
        "venta_abastecimiento.html",
        {
            "ventas": ventas,
            "canales": canales,
            "form": form,
            "edit_venta": None,
            "delete_error": None,
            "import_error": None,
        },
    )


def venta_update(request, pk):
    if get_user_from_request(request) is None:
        return redirect("login")
    venta = get_object_or_404(VentaAbastecimiento, pk=pk)
    if request.method == "POST":
        form = VentaForm(request.POST, instance=venta)
        if form.is_valid():
            form.save()
            return redirect("abastecimientos:venta_list")
    else:
        form = VentaForm(instance=venta)
    ventas = (
        VentaAbastecimiento.objects.select_related(
            "id_producto", "id_puntoventa", "id_canal"
        )
        .order_by("-fecha_venta", "id_venta")
    )
    canales = Canal.objects.order_by("canal_nombre")
    return render(
        request,
        "venta_abastecimiento.html",
        {
            "ventas": ventas,
            "canales": canales,
            "form": form,
            "edit_venta": venta,
            "delete_error": None,
            "import_error": None,
        },
    )


def venta_delete(request, pk):
    if get_user_from_request(request) is None:
        return redirect("login")
    venta = get_object_or_404(VentaAbastecimiento, pk=pk)
    if request.method == "POST":
        try:
            venta.delete()
            return redirect("abastecimientos:venta_list")
        except ProtectedError:
            ventas = (
                VentaAbastecimiento.objects.select_related(
                    "id_producto", "id_puntoventa", "id_canal"
                )
                .order_by("-fecha_venta", "id_venta")
            )
            canales = Canal.objects.order_by("canal_nombre")
            form = VentaForm()
            return render(
                request,
                "venta_abastecimiento.html",
                {
                    "ventas": ventas,
                    "canales": canales,
                    "form": form,
                    "edit_venta": None,
                    "delete_error": "No se puede eliminar: tiene registros relacionados.",
                    "import_error": None,
                },
            )
    return redirect("abastecimientos:venta_list")


def venta_export(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    queryset = VentaAbastecimiento.objects.order_by("id_venta")
    return _build_xlsx_response(
        queryset, "venta_abastecimiento.xlsx", "Ventas", exclude_fields=["id_canal"]
    )

from django.db import transaction

def venta_import(request):
    if get_user_from_request(request) is None:
        return redirect("login")

    if request.method != "POST":
        return redirect("abastecimientos:venta_list")

    archivo = request.FILES.get("archivo")
    if not archivo:
        return redirect("abastecimientos:venta_list")

    try:
        headers, rows = _load_xlsx_rows(archivo)

        required = [
            "id_producto",
            "id_puntoventa",
            "cantidad_venta",
            "fecha_venta",
        ]
        for name in required:
            if name not in headers:
                raise ValueError(f"Falta la columna {name}.")

        canal_id = request.POST.get("canal_id")
        if canal_id:
            canal_id = str(canal_id).strip()

        has_id_canal = "id_canal" in headers
        if not canal_id and not has_id_canal:
            raise ValueError("Seleccione un canal o incluya la columna id_canal.")

        header_index = {name: idx for idx, name in enumerate(headers)}

        # Cargar validaciones una sola vez
        productos_validos = set(ProductoAbastecimiento.objects.values_list("pk", flat=True))
        puntoventas_validos = set(PuntoVentaAbastecimiento.objects.values_list("pk", flat=True))
        canales_validos = set(Canal.objects.values_list("pk", flat=True))

        ventas_a_crear = []
        ventas_a_actualizar = []

        # Buscar ventas existentes por id_venta solo si vienen en el archivo
        ids_existentes = [
            int(_get_cell(row, header_index["id_venta"]))
            for row in rows
            if "id_venta" in header_index
            and _get_cell(row, header_index["id_venta"]) not in [None, ""]
        ]

        ventas_existentes = {
            v.id_venta: v
            for v in VentaAbastecimiento.objects.filter(id_venta__in=ids_existentes)
        }

        with transaction.atomic():
            for row in rows:
                if _row_is_empty(row):
                    continue

                id_venta = (
                    _get_cell(row, header_index["id_venta"])
                    if "id_venta" in header_index
                    else None
                )

                id_producto = _get_cell(row, header_index["id_producto"])
                id_puntoventa = _get_cell(row, header_index["id_puntoventa"])
                cantidad = _get_cell(row, header_index["cantidad_venta"])
                fecha = _get_cell(row, header_index["fecha_venta"])
                id_canal_row = canal_id if canal_id else _get_cell(row, header_index.get("id_canal"))

                id_producto = "" if id_producto is None else str(id_producto).strip()
                id_puntoventa = "" if id_puntoventa is None else str(id_puntoventa).strip()
                id_canal_row = "" if id_canal_row is None else str(id_canal_row).strip()

                if not id_producto or not id_puntoventa or not id_canal_row:
                    raise ValueError("Faltan campos requeridos en ventas.")

                if cantidad is None or str(cantidad).strip() == "":
                    raise ValueError("cantidad_venta es requerida.")

                cantidad = int(cantidad)
                fecha = _parse_date(fecha)
                if fecha is None:
                    raise ValueError("fecha_venta es requerida.")

                if id_producto not in productos_validos:
                    raise ValueError(f"Producto no existe: {id_producto}")

                if id_puntoventa not in puntoventas_validos:
                    raise ValueError(f"Punto de venta no existe: {id_puntoventa}")

                if id_canal_row not in canales_validos:
                    raise ValueError(f"Canal no existe: {id_canal_row}")

                if id_venta is not None and str(id_venta).strip() != "":
                    id_venta = int(id_venta)

                    if id_venta in ventas_existentes:
                        venta = ventas_existentes[id_venta]
                        venta.id_producto_id = id_producto
                        venta.id_puntoventa_id = id_puntoventa
                        venta.cantidad_venta = cantidad
                        venta.fecha_venta = fecha
                        venta.id_canal_id = id_canal_row
                        ventas_a_actualizar.append(venta)
                    else:
                        ventas_a_crear.append(
                            VentaAbastecimiento(
                                id_producto_id=id_producto,
                                id_puntoventa_id=id_puntoventa,
                                cantidad_venta=cantidad,
                                fecha_venta=fecha,
                                id_canal_id=id_canal_row,
                            )
                        )
                else:
                    ventas_a_crear.append(
                        VentaAbastecimiento(
                            id_producto_id=id_producto,
                            id_puntoventa_id=id_puntoventa,
                            cantidad_venta=cantidad,
                            fecha_venta=fecha,
                            id_canal_id=id_canal_row,
                        )
                    )

            if ventas_a_crear:
                VentaAbastecimiento.objects.bulk_create(ventas_a_crear, batch_size=500)

            if ventas_a_actualizar:
                VentaAbastecimiento.objects.bulk_update(
                    ventas_a_actualizar,
                    ["id_producto", "id_puntoventa", "cantidad_venta", "fecha_venta", "id_canal"],
                    batch_size=500,
                )

        return redirect("abastecimientos:venta_list")

    except Exception as exc:
        ventas = (
            VentaAbastecimiento.objects.select_related(
                "id_producto", "id_puntoventa", "id_canal"
            )
            .order_by("-fecha_venta", "id_venta")
        )
        canales = Canal.objects.order_by("canal_nombre")
        form = VentaForm()
        return render(
            request,
            "venta_abastecimiento.html",
            {
                "ventas": ventas,
                "canales": canales,
                "form": form,
                "edit_venta": None,
                "delete_error": None,
                "import_error": f"Error al importar: {exc}",
            },
        )

def venta_delete_by_canal(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    if request.method != "POST":
        return redirect("abastecimientos:venta_list")
    canal_id = request.POST.get("canal_id")
    if canal_id:
        canal_id = str(canal_id).strip()
    if canal_id:
        VentaAbastecimiento.objects.filter(id_canal_id=canal_id).delete()
    return redirect("abastecimientos:venta_list")


def _normalize_str(value):
    if value is None:
        return ""
    return str(value).strip()


def _parse_multi_param(request, name):
    value = request.GET.get(name)
    if not value:
        return []
    parts = [p.strip() for p in value.split(",") if p.strip() != ""]
    return [p.lower() for p in parts]


def _build_cruzar_records():
    def to_int(value):
        if value is None or value == "":
            return 0
        try:
            return int(value)
        except (TypeError, ValueError):
            try:
                return int(float(value))
            except (TypeError, ValueError):
                return 0

    registros = {}

    for row in AbastecimientoClaro.objects.values(
        "material",
        "producto",
        "centro_costos",
        "nombre_punto",
        "inventario_claro",
        "transito_claro",
        "ventas_actuales_claro",
        "sugerido_claro",
    ):
        material = _normalize_str(row.get("material"))
        centro_costos = _normalize_str(row.get("centro_costos"))
        if not material or not centro_costos:
            continue
        key = f"{material}|{centro_costos}"
        registros[key] = {
            "Material": material,
            "Producto": _normalize_str(row.get("producto")),
            "Marca": "",
            "Centro Costos": centro_costos,
            "Punto de Venta": _normalize_str(row.get("nombre_punto")),
            "Sugerido Claro": to_int(row.get("sugerido_claro")),
            "Inventario": to_int(row.get("inventario_claro")),
            "Transitos": to_int(row.get("transito_claro")),
            "Ventas Actuales": to_int(row.get("ventas_actuales_claro")),
            "Envio Inventario 3 meses": 0,
            "Sugerido Coltrade": 0,
            "Promedio 3 Meses": 0,
            "Sugerido Final": 0,
        }

    for row in AbastecimientoColtrade.objects.values(
        "material",
        "producto",
        "marca",
        "centro_costos",
        "punto_venta",
        "ventas_actuales",
        "transitos",
        "inventario",
        "envio_inventario_3_meses",
        "sugerido_coltrade",
    ):
        material = _normalize_str(row.get("material"))
        centro_costos = _normalize_str(row.get("centro_costos"))
        if not material or not centro_costos:
            continue
        key = f"{material}|{centro_costos}"
        if key not in registros:
            registros[key] = {
                "Material": material,
                "Producto": _normalize_str(row.get("producto")),
                "Marca": _normalize_str(row.get("marca")),
                "Centro Costos": centro_costos,
                "Punto de Venta": _normalize_str(row.get("punto_venta")),
                "Sugerido Claro": 0,
                "Inventario": to_int(row.get("inventario")),
                "Transitos": to_int(row.get("transitos")),
                "Ventas Actuales": to_int(row.get("ventas_actuales")),
                "Envio Inventario 3 meses": to_int(row.get("envio_inventario_3_meses")),
                "Sugerido Coltrade": to_int(row.get("sugerido_coltrade")),
                "Promedio 3 Meses": 0,
                "Sugerido Final": 0,
            }
        else:
            registro = registros[key]
            if not registro.get("Producto"):
                registro["Producto"] = _normalize_str(row.get("producto"))
            if not registro.get("Marca"):
                registro["Marca"] = _normalize_str(row.get("marca"))
            if not registro.get("Punto de Venta"):
                registro["Punto de Venta"] = _normalize_str(row.get("punto_venta"))

            registro["Sugerido Coltrade"] = to_int(row.get("sugerido_coltrade"))
            registro["Envio Inventario 3 meses"] = to_int(
                row.get("envio_inventario_3_meses")
            )

            if registro.get("Inventario", 0) == 0 and row.get("inventario"):
                registro["Inventario"] = to_int(row.get("inventario"))
            if registro.get("Transitos", 0) == 0 and row.get("transitos"):
                registro["Transitos"] = to_int(row.get("transitos"))
            if registro.get("Ventas Actuales", 0) == 0 and row.get("ventas_actuales"):
                registro["Ventas Actuales"] = to_int(row.get("ventas_actuales"))

    records = sorted(
        registros.values(),
        key=lambda x: (x.get("Centro Costos") or "", x.get("Material") or ""),
    )
    return records


def cruzar_archivos_page(request):
    user = get_user_from_request(request)
    if user is None:
        return redirect("login")
    return render(request, "cruzar_archivos.html", {"jwt_user": user})


def cruzar_archivos_data(request):
    user = get_user_from_request(request)
    if user is None:
        return redirect("login")
    try:
        records = _build_cruzar_records()
        return JsonResponse({"status": "ok", "data": records})
    except Exception as exc:
        return JsonResponse({"status": "error", "message": str(exc)}, status=500)


def cruzar_archivos_export(request):
    user = get_user_from_request(request)
    if user is None:
        return redirect("login")

    records = _build_cruzar_records()
    headers = [
        "Material",
        "Producto",
        "Marca",
        "Centro Costos",
        "Punto de Venta",
        "Sugerido Claro",
        "Inventario",
        "Transitos",
        "Ventas Actuales",
        "Envio Inventario 3 meses",
        "Sugerido Coltrade",
        "Promedio 3 Meses",
        "Sugerido Final",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Cruzado"
    ws.append(headers)
    for row in records:
        ws.append([row.get(h, "") for h in headers])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=cruzar_archivos.xlsx"
    return response


MARCAS_ESPECIFICAS = [
    "Aiwa",
    "Amazon",
    "Belkin",
    "Pmp",
    "Haxly",
    "Sylvania",
    "Roku",
    "Nintendo",
    "Redragon",
    "Spigen",
    "Motorola",
    "Logitech",
    "Zte",
    "Cubitt",
]


def _increment_serial(serial):
    if serial is None or str(serial).strip() == "":
        return ""
    serial_str = str(serial)
    matches = list(re.finditer(r"(\d+)", serial_str))
    if not matches:
        return serial_str + "1"
    last_match = matches[-1]
    start, end = last_match.span()
    num = int(last_match.group())
    incremented_num = num + 1
    return serial_str[:start] + str(incremented_num) + serial_str[end:]


def _to_int(value):
    if value is None or str(value).strip() == "":
        return 0
    try:
        return int(value)
    except (TypeError, ValueError):
        try:
            return int(float(str(value).strip()))
        except (TypeError, ValueError):
            return 0


def serializar_ventas_page(request):
    user = get_user_from_request(request)
    if user is None:
        return redirect("login")
    return render(
        request,
        "serializar_ventas.html",
        {"jwt_user": user, "marcas_especificas": MARCAS_ESPECIFICAS},
    )


def serializar_ventas_preview(request):
    user = get_user_from_request(request)
    if user is None:
        return redirect("login")

    if request.method != "POST":
        return JsonResponse({"error": "M??todo no permitido"}, status=405)

    archivo = request.FILES.get("file")
    if not archivo:
        return JsonResponse({"error": "No se envi?? archivo"}, status=400)

    try:
        headers, rows = _load_xlsx_rows(archivo)
    except Exception as exc:
        return JsonResponse({"error": f"Error al leer el archivo: {exc}"}, status=400)

    header_index = {name: idx for idx, name in enumerate(headers)}

    marcas = set()
    if "Marca" in header_index and "Sugerido Final" in header_index:
        for row in rows:
            if _row_is_empty(row):
                continue
            sugerido_val = _get_cell(row, header_index["Sugerido Final"])
            sugerido = _to_int(sugerido_val)
            if sugerido == 0:
                continue
            marca = _get_cell(row, header_index["Marca"])
            marca = _normalize_str(marca)
            if marca:
                marcas.add(marca)

    return JsonResponse({"marcas": sorted(list(marcas))})


def serializar_ventas_process(request):
    user = get_user_from_request(request)
    if user is None:
        return redirect("login")

    if request.method != "POST":
        return HttpResponse("M??todo no permitido", status=405)

    archivo = request.FILES.get("file")
    if not archivo:
        return HttpResponse("No se ha proporcionado un archivo", status=400)

    try:
        headers, rows = _load_xlsx_rows(archivo)
    except Exception as exc:
        return HttpResponse(f"Error al leer el archivo: {exc}", status=400)

    required_columns = [
        "Material",
        "Producto",
        "Marca",
        "Centro Costos",
        "Punto de Venta",
        "Sugerido Final",
    ]
    missing_cols = [col for col in required_columns if col not in headers]
    if missing_cols:
        return HttpResponse(
            f"Faltan columnas requeridas: {', '.join(missing_cols)}", status=400
        )

    header_index = {name: idx for idx, name in enumerate(headers)}

    seriales_iniciales = {}
    for marca in MARCAS_ESPECIFICAS:
        key = marca.lower()
        serial_val = _normalize_str(request.POST.get(f"serial_{key}"))
        if key == "haxly" and serial_val == "250":
            serial_val = "201"
        seriales_iniciales[key] = serial_val

    otros_serial = _normalize_str(request.POST.get("otros_serial"))
    ultimos_seriales = {}

    grupos = {}
    for row in rows:
        if _row_is_empty(row):
            continue
        sugerido_val = _get_cell(row, header_index["Sugerido Final"])
        sugerido = _to_int(sugerido_val)
        if sugerido == 0:
            continue

        centro = _normalize_str(_get_cell(row, header_index["Centro Costos"]))
        punto = _normalize_str(_get_cell(row, header_index["Punto de Venta"]))
        material = _normalize_str(_get_cell(row, header_index["Material"]))
        producto = _normalize_str(_get_cell(row, header_index["Producto"]))
        marca = _normalize_str(_get_cell(row, header_index["Marca"]))

        key = (centro, punto, material, producto, marca, sugerido)
        grupos[key] = True

    final_rows = []
    today_str = datetime.today().strftime("%Y-%m-%d")

    for (centro, punto, material, producto, marca, sugerido) in grupos.keys():
        num_registros = int(sugerido)
        if num_registros <= 0:
            continue

        marca_key = marca.lower()
        serial_asignado = None

        for marca_esp in MARCAS_ESPECIFICAS:
            if marca_esp.lower() in marca_key:
                marca_key = marca_esp.lower()
                if seriales_iniciales.get(marca_key):
                    serial_asignado = seriales_iniciales[marca_key]
                break
        else:
            if otros_serial:
                marca_key = "otros"
                serial_asignado = otros_serial

        serial_actual = serial_asignado
        for i in range(num_registros):
            serial_val = ""
            if serial_actual:
                serial_val = serial_actual
                ultimos_seriales[marca] = serial_actual
                serial_actual = _increment_serial(serial_actual)

            final_rows.append(
                {
                    "No": i + 1,
                    "Centro Costos": centro,
                    "Punto de Venta": punto,
                    "Material": material,
                    "Producto": producto,
                    "Marca": marca,
                    "Fecha Actual": today_str,
                    "Serial": serial_val,
                    "Sugerido Final": sugerido,
                }
            )

        if serial_asignado:
            if marca_key in seriales_iniciales:
                seriales_iniciales[marca_key] = serial_actual
            elif marca_key == "otros":
                otros_serial = serial_actual

        final_rows.append(
            {
                "No": "",
                "Centro Costos": "",
                "Punto de Venta": "Recuento de Unidades",
                "Material": num_registros,
                "Producto": "",
                "Marca": "",
                "Fecha Actual": "",
                "Serial": "",
                "Sugerido Final": "",
            }
        )

    if ultimos_seriales:
        resumen_rows = []
        for marca, serial in ultimos_seriales.items():
            registro = "No"
            for marca_esp in MARCAS_ESPECIFICAS:
                if marca_esp.lower() in marca.lower():
                    registro = "Si"
                    break
            resumen_rows.append(
                {"Marca": marca, "Ultimo Serial": serial, "??Registro?": registro}
            )
        resumen_rows = sorted(resumen_rows, key=lambda x: x.get("Marca") or "")
    else:
        resumen_rows = []

    headers_out = [
        "No",
        "Centro Costos",
        "Punto de Venta",
        "Material",
        "Producto",
        "Marca",
        "Fecha Actual",
        "Serial",
        "Sugerido Final",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Serializado"
    ws.append(headers_out)

    bold_font = Font(bold=True)
    for row in final_rows:
        ws.append([row.get(h, "") for h in headers_out])
        if row.get("Punto de Venta") == "Recuento de Unidades":
            row_idx = ws.max_row
            for col in range(1, len(headers_out) + 1):
                ws.cell(row=row_idx, column=col).font = bold_font

    ws_resumen = wb.create_sheet("Resumen Seriales")
    resumen_headers = ["Marca", "Ultimo Serial", "??Registro?"]
    ws_resumen.append(resumen_headers)
    for row in resumen_rows:
        ws_resumen.append([row.get(h, "") for h in resumen_headers])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=archivo_serializado.xlsx"
    return response


def _parse_uuid(value):
    if value is None or str(value).strip() == "":
        return None
    try:
        return uuid.UUID(str(value).strip())
    except Exception as exc:
        raise ValueError(f"UUID invalido: {value}") from exc


def _safe_int(value):
    if value is None or str(value).strip() == "":
        return 0
    try:
        return int(value)
    except (TypeError, ValueError):
        try:
            return int(float(str(value).strip()))
        except (TypeError, ValueError):
            return 0


def abastecimiento_claro_list(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    claros = AbastecimientoClaro.objects.order_by("centro_costos", "material")
    form = AbastecimientoClaroForm()
    return render(
        request,
        "data_abastecimiento_claro.html",
        {
            "claros": claros,
            "form": form,
            "edit_claro": None,
            "delete_error": None,
            "import_error": None,
        },
    )


def abastecimiento_claro_create(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    if request.method != "POST":
        return redirect("abastecimientos:abastecimiento_claro_list")
    form = AbastecimientoClaroForm(request.POST)
    if form.is_valid():
        form.save()
        return redirect("abastecimientos:abastecimiento_claro_list")
    claros = AbastecimientoClaro.objects.order_by("centro_costos", "material")
    return render(
        request,
        "data_abastecimiento_claro.html",
        {
            "claros": claros,
            "form": form,
            "edit_claro": None,
            "delete_error": None,
            "import_error": None,
        },
    )


def abastecimiento_claro_update(request, pk):
    if get_user_from_request(request) is None:
        return redirect("login")
    claro = get_object_or_404(AbastecimientoClaro, pk=pk)
    if request.method == "POST":
        form = AbastecimientoClaroForm(request.POST, instance=claro)
        if form.is_valid():
            form.save()
            return redirect("abastecimientos:abastecimiento_claro_list")
    else:
        form = AbastecimientoClaroForm(instance=claro)
    claros = AbastecimientoClaro.objects.order_by("centro_costos", "material")
    return render(
        request,
        "data_abastecimiento_claro.html",
        {
            "claros": claros,
            "form": form,
            "edit_claro": claro,
            "delete_error": None,
            "import_error": None,
        },
    )


def abastecimiento_claro_delete(request, pk):
    if get_user_from_request(request) is None:
        return redirect("login")
    claro = get_object_or_404(AbastecimientoClaro, pk=pk)
    if request.method == "POST":
        try:
            claro.delete()
            return redirect("abastecimientos:abastecimiento_claro_list")
        except ProtectedError:
            claros = AbastecimientoClaro.objects.order_by("centro_costos", "material")
            form = AbastecimientoClaroForm()
            return render(
                request,
                "data_abastecimiento_claro.html",
                {
                    "claros": claros,
                    "form": form,
                    "edit_claro": None,
                    "delete_error": "No se puede eliminar: tiene registros relacionados.",
                    "import_error": None,
                },
            )
    return redirect("abastecimientos:abastecimiento_claro_list")


def abastecimiento_claro_export(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    queryset = AbastecimientoClaro.objects.order_by("centro_costos", "material")
    return _build_xlsx_response(queryset, "abastecimiento_claro.xlsx", "Claro")


def abastecimiento_claro_import(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    if request.method != "POST":
        return redirect("abastecimientos:abastecimiento_claro_list")
    archivo = request.FILES.get("archivo")
    if not archivo:
        return redirect("abastecimientos:abastecimiento_claro_list")
    try:
        headers, rows = _load_xlsx_rows(archivo)
        required = [
            "material",
            "producto",
            "centro_costos",
            "nombre_punto",
            "inventario_claro",
            "transito_claro",
            "ventas_pasadas_claro",
            "ventas_actuales_claro",
            "sugerido_claro",
        ]
        for name in required:
            if name not in headers:
                raise ValueError(f"Falta la columna {name}.")

        header_index = {name: idx for idx, name in enumerate(headers)}

        ids_existentes = []
        if "id" in header_index:
            for row in rows:
                if _row_is_empty(row):
                    continue
                id_val = _get_cell(row, header_index["id"])
                if id_val is None or str(id_val).strip() == "":
                    continue
                ids_existentes.append(_parse_uuid(id_val))

        claros_existentes = {}
        if ids_existentes:
            claros_existentes = {
                obj.id: obj for obj in AbastecimientoClaro.objects.filter(id__in=ids_existentes)
            }

        claros_a_crear = []
        claros_a_actualizar = []

        with transaction.atomic():
            for row in rows:
                if _row_is_empty(row):
                    continue

                id_val = None
                if "id" in header_index:
                    raw_id = _get_cell(row, header_index["id"])
                    if raw_id is not None and str(raw_id).strip() != "":
                        id_val = _parse_uuid(raw_id)

                material = _normalize_str(_get_cell(row, header_index["material"]))
                producto = _normalize_str(_get_cell(row, header_index["producto"]))
                centro_costos = _normalize_str(_get_cell(row, header_index["centro_costos"]))
                nombre_punto = _normalize_str(_get_cell(row, header_index["nombre_punto"]))
                inventario = _safe_int(_get_cell(row, header_index["inventario_claro"]))
                transito = _safe_int(_get_cell(row, header_index["transito_claro"]))
                ventas_pasadas = _safe_int(_get_cell(row, header_index["ventas_pasadas_claro"]))
                ventas_actuales = _safe_int(_get_cell(row, header_index["ventas_actuales_claro"]))
                sugerido = _safe_int(_get_cell(row, header_index["sugerido_claro"]))

                if not material or not producto or not centro_costos or not nombre_punto:
                    raise ValueError("Faltan campos requeridos en abastecimiento claro.")

                if id_val and id_val in claros_existentes:
                    claro = claros_existentes[id_val]
                    claro.material = material
                    claro.producto = producto
                    claro.centro_costos = centro_costos
                    claro.nombre_punto = nombre_punto
                    claro.inventario_claro = inventario
                    claro.transito_claro = transito
                    claro.ventas_pasadas_claro = ventas_pasadas
                    claro.ventas_actuales_claro = ventas_actuales
                    claro.sugerido_claro = sugerido
                    claros_a_actualizar.append(claro)
                else:
                    claros_a_crear.append(
                        AbastecimientoClaro(
                            id=id_val or uuid.uuid4(),
                            material=material,
                            producto=producto,
                            centro_costos=centro_costos,
                            nombre_punto=nombre_punto,
                            inventario_claro=inventario,
                            transito_claro=transito,
                            ventas_pasadas_claro=ventas_pasadas,
                            ventas_actuales_claro=ventas_actuales,
                            sugerido_claro=sugerido,
                        )
                    )

            if claros_a_crear:
                AbastecimientoClaro.objects.bulk_create(claros_a_crear, batch_size=500)
            if claros_a_actualizar:
                AbastecimientoClaro.objects.bulk_update(
                    claros_a_actualizar,
                    [
                        "material",
                        "producto",
                        "centro_costos",
                        "nombre_punto",
                        "inventario_claro",
                        "transito_claro",
                        "ventas_pasadas_claro",
                        "ventas_actuales_claro",
                        "sugerido_claro",
                    ],
                    batch_size=500,
                )

        return redirect("abastecimientos:abastecimiento_claro_list")
    except Exception as exc:
        claros = AbastecimientoClaro.objects.order_by("centro_costos", "material")
        form = AbastecimientoClaroForm()
        return render(
            request,
            "data_abastecimiento_claro.html",
            {
                "claros": claros,
                "form": form,
                "edit_claro": None,
                "delete_error": None,
                "import_error": f"Error al importar: {exc}",
            },
        )


def abastecimiento_coltrade_list(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    coltrade = AbastecimientoColtrade.objects.order_by("centro_costos", "material")
    form = AbastecimientoColtradeForm()
    return render(
        request,
        "data_abastecimiento_coltrade.html",
        {
            "coltrade": coltrade,
            "form": form,
            "edit_coltrade": None,
            "delete_error": None,
            "import_error": None,
        },
    )


def abastecimiento_coltrade_create(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    if request.method != "POST":
        return redirect("abastecimientos:abastecimiento_coltrade_list")
    form = AbastecimientoColtradeForm(request.POST)
    if form.is_valid():
        form.save()
        return redirect("abastecimientos:abastecimiento_coltrade_list")
    coltrade = AbastecimientoColtrade.objects.order_by("centro_costos", "material")
    return render(
        request,
        "data_abastecimiento_coltrade.html",
        {
            "coltrade": coltrade,
            "form": form,
            "edit_coltrade": None,
            "delete_error": None,
            "import_error": None,
        },
    )


def abastecimiento_coltrade_update(request, pk):
    if get_user_from_request(request) is None:
        return redirect("login")
    registro = get_object_or_404(AbastecimientoColtrade, pk=pk)
    if request.method == "POST":
        form = AbastecimientoColtradeForm(request.POST, instance=registro)
        if form.is_valid():
            form.save()
            return redirect("abastecimientos:abastecimiento_coltrade_list")
    else:
        form = AbastecimientoColtradeForm(instance=registro)
    coltrade = AbastecimientoColtrade.objects.order_by("centro_costos", "material")
    return render(
        request,
        "data_abastecimiento_coltrade.html",
        {
            "coltrade": coltrade,
            "form": form,
            "edit_coltrade": registro,
            "delete_error": None,
            "import_error": None,
        },
    )


def abastecimiento_coltrade_delete(request, pk):
    if get_user_from_request(request) is None:
        return redirect("login")
    registro = get_object_or_404(AbastecimientoColtrade, pk=pk)
    if request.method == "POST":
        try:
            registro.delete()
            return redirect("abastecimientos:abastecimiento_coltrade_list")
        except ProtectedError:
            coltrade = AbastecimientoColtrade.objects.order_by("centro_costos", "material")
            form = AbastecimientoColtradeForm()
            return render(
                request,
                "data_abastecimiento_coltrade.html",
                {
                    "coltrade": coltrade,
                    "form": form,
                    "edit_coltrade": None,
                    "delete_error": "No se puede eliminar: tiene registros relacionados.",
                    "import_error": None,
                },
            )
    return redirect("abastecimientos:abastecimiento_coltrade_list")


def abastecimiento_coltrade_export(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    queryset = AbastecimientoColtrade.objects.order_by("centro_costos", "material")
    return _build_xlsx_response(queryset, "abastecimiento_coltrade.xlsx", "Coltrade")


def abastecimiento_coltrade_import(request):
    if get_user_from_request(request) is None:
        return redirect("login")
    if request.method != "POST":
        return redirect("abastecimientos:abastecimiento_coltrade_list")
    archivo = request.FILES.get("archivo")
    if not archivo:
        return redirect("abastecimientos:abastecimiento_coltrade_list")
    try:
        headers, rows = _load_xlsx_rows(archivo)
        required = [
            "centro_costos",
            "punto_venta",
            "material",
            "producto",
            "marca",
            "ventas_actuales",
            "transitos",
            "inventario",
            "envio_inventario_3_meses",
            "sugerido_coltrade",
        ]
        for name in required:
            if name not in headers:
                raise ValueError(f"Falta la columna {name}.")

        header_index = {name: idx for idx, name in enumerate(headers)}

        ids_existentes = []
        if "id" in header_index:
            for row in rows:
                if _row_is_empty(row):
                    continue
                id_val = _get_cell(row, header_index["id"])
                if id_val is None or str(id_val).strip() == "":
                    continue
                ids_existentes.append(_parse_uuid(id_val))

        coltrade_existentes = {}
        if ids_existentes:
            coltrade_existentes = {
                obj.id: obj
                for obj in AbastecimientoColtrade.objects.filter(id__in=ids_existentes)
            }

        coltrade_a_crear = []
        coltrade_a_actualizar = []

        with transaction.atomic():
            for row in rows:
                if _row_is_empty(row):
                    continue

                id_val = None
                if "id" in header_index:
                    raw_id = _get_cell(row, header_index["id"])
                    if raw_id is not None and str(raw_id).strip() != "":
                        id_val = _parse_uuid(raw_id)

                centro_costos = _normalize_str(_get_cell(row, header_index["centro_costos"]))
                punto_venta = _normalize_str(_get_cell(row, header_index["punto_venta"]))
                material = _normalize_str(_get_cell(row, header_index["material"]))
                producto = _normalize_str(_get_cell(row, header_index["producto"]))
                marca = _normalize_str(_get_cell(row, header_index["marca"]))
                ventas_actuales = _safe_int(_get_cell(row, header_index["ventas_actuales"]))
                transitos = _safe_int(_get_cell(row, header_index["transitos"]))
                inventario = _safe_int(_get_cell(row, header_index["inventario"]))
                envio_3 = _safe_int(_get_cell(row, header_index["envio_inventario_3_meses"]))
                sugerido = _safe_int(_get_cell(row, header_index["sugerido_coltrade"]))

                if not centro_costos or not punto_venta or not material or not producto:
                    raise ValueError("Faltan campos requeridos en abastecimiento coltrade.")

                if id_val and id_val in coltrade_existentes:
                    registro = coltrade_existentes[id_val]
                    registro.centro_costos = centro_costos
                    registro.punto_venta = punto_venta
                    registro.material = material
                    registro.producto = producto
                    registro.marca = marca
                    registro.ventas_actuales = ventas_actuales
                    registro.transitos = transitos
                    registro.inventario = inventario
                    registro.envio_inventario_3_meses = envio_3
                    registro.sugerido_coltrade = sugerido
                    coltrade_a_actualizar.append(registro)
                else:
                    coltrade_a_crear.append(
                        AbastecimientoColtrade(
                            id=id_val or uuid.uuid4(),
                            centro_costos=centro_costos,
                            punto_venta=punto_venta,
                            material=material,
                            producto=producto,
                            marca=marca,
                            ventas_actuales=ventas_actuales,
                            transitos=transitos,
                            inventario=inventario,
                            envio_inventario_3_meses=envio_3,
                            sugerido_coltrade=sugerido,
                        )
                    )

            if coltrade_a_crear:
                AbastecimientoColtrade.objects.bulk_create(coltrade_a_crear, batch_size=500)
            if coltrade_a_actualizar:
                AbastecimientoColtrade.objects.bulk_update(
                    coltrade_a_actualizar,
                    [
                        "centro_costos",
                        "punto_venta",
                        "material",
                        "producto",
                        "marca",
                        "ventas_actuales",
                        "transitos",
                        "inventario",
                        "envio_inventario_3_meses",
                        "sugerido_coltrade",
                    ],
                    batch_size=500,
                )

        return redirect("abastecimientos:abastecimiento_coltrade_list")
    except Exception as exc:
        coltrade = AbastecimientoColtrade.objects.order_by("centro_costos", "material")
        form = AbastecimientoColtradeForm()
        return render(
            request,
            "data_abastecimiento_coltrade.html",
            {
                "coltrade": coltrade,
                "form": form,
                "edit_coltrade": None,
                "delete_error": None,
                "import_error": f"Error al importar: {exc}",
            },
        )


def forecast_page(request):
    user = get_user_from_request(request)
    if user is None:
        return redirect("login")
    return render(request, "forecast.html", {"jwt_user": user})


def forecast_options(request):
    user = get_user_from_request(request)
    if user is None:
        return redirect("login")

    id_puntoventa_filter = _parse_multi_param(request, "id_puntoventa")
    punto_venta_filter = _parse_multi_param(request, "punto_venta")
    id_producto_filter = _parse_multi_param(request, "id_producto")
    nombre_producto_filter = _parse_multi_param(request, "nombre_producto")
    marcas_filter = _parse_multi_param(request, "marca")
    canal_regional_filter = _parse_multi_param(request, "canal_regional")

    centros = set()
    puntos = set()
    materiales = set()
    productos = set()
    marcas = set()
    canales = set()

    for punto in PuntoVentaAbastecimiento.objects.all():
        cc = _normalize_str(punto.id_puntoventa)
        pv = _normalize_str(punto.punto_venta)
        canal = _normalize_str(punto.canal_regional)
        if id_puntoventa_filter and cc.lower() not in id_puntoventa_filter:
            continue
        if punto_venta_filter and pv.lower() not in punto_venta_filter:
            continue
        if canal_regional_filter and canal.lower() not in canal_regional_filter:
            continue
        centros.add(cc)
        puntos.add(pv)
        if canal:
            canales.add(canal)

    for cc in (
        InventarioAbastecimiento.objects.values_list(
            "id_puntoventa_id", flat=True
        ).distinct()
    ):
        cc = _normalize_str(cc)
        if id_puntoventa_filter and cc.lower() not in id_puntoventa_filter:
            continue
        centros.add(cc)

    for cc in (
        TransitoAbastecimiento.objects.values_list(
            "id_puntoventa_id", flat=True
        ).distinct()
    ):
        cc = _normalize_str(cc)
        if id_puntoventa_filter and cc.lower() not in id_puntoventa_filter:
            continue
        centros.add(cc)

    for cc in (
        VentaAbastecimiento.objects.values_list("id_puntoventa_id", flat=True).distinct()
    ):
        cc = _normalize_str(cc)
        if id_puntoventa_filter and cc.lower() not in id_puntoventa_filter:
            continue
        centros.add(cc)

    for prod in ProductoAbastecimiento.objects.all():
        mat = _normalize_str(prod.id_producto)
        prod_name = _normalize_str(prod.nombre_producto)
        marca = _normalize_str(prod.marca)
        if id_producto_filter and mat.lower() not in id_producto_filter:
            continue
        if nombre_producto_filter and prod_name and prod_name.lower() not in nombre_producto_filter:
            continue
        if marcas_filter and marca != "" and marca.lower() not in marcas_filter:
            continue
        materiales.add(mat)
        if prod_name:
            productos.add(prod_name)
        if marca:
            marcas.add(marca)

    return JsonResponse(
        {
            "id_puntoventa": sorted(list(centros)),
            "punto_venta": sorted(list(puntos)),
            "id_producto": sorted(list(materiales)),
            "nombre_producto": sorted(list(productos)),
            "marcas": sorted(list(marcas)),
            "canal_regional": sorted(list(canales)),
        }
    )


def forecast_data(request):
    user = get_user_from_request(request)
    if user is None:
        return redirect("login")

    id_puntoventa_filter = _parse_multi_param(request, "id_puntoventa")
    punto_venta_filter = _parse_multi_param(request, "punto_venta")
    id_producto_filter = _parse_multi_param(request, "id_producto")
    nombre_producto_filter = _parse_multi_param(request, "nombre_producto")
    marcas_filter = _parse_multi_param(request, "marca")
    canal_regional_filter = _parse_multi_param(request, "canal_regional")

    try:
        page = max(1, int(request.GET.get("page", 1)))
    except Exception:
        page = 1
    try:
        page_size = int(request.GET.get("page_size", 50))
        if page_size <= 0:
            page_size = 50
    except Exception:
        page_size = 50
    if page_size > 1000:
        page_size = 1000

    prod_map = {}
    for prod in ProductoAbastecimiento.objects.all():
        mat = _normalize_str(prod.id_producto)
        prod_map[mat] = {
            "Producto": _normalize_str(prod.nombre_producto),
            "Marca": _normalize_str(prod.marca),
        }

    puntos_map = {}
    for punto in PuntoVentaAbastecimiento.objects.all():
        cc = _normalize_str(punto.id_puntoventa)
        puntos_map[cc] = {
            "Punto de Venta": _normalize_str(punto.punto_venta),
            "Canal o Regional": _normalize_str(punto.canal_regional),
        }

    candidates = set()
    for cc, mat in InventarioAbastecimiento.objects.values_list(
        "id_puntoventa_id", "id_producto_id"
    ).distinct():
        cc = _normalize_str(cc)
        mat = _normalize_str(mat)
        if mat in prod_map:
            candidates.add((cc, mat))

    for cc, mat in TransitoAbastecimiento.objects.values_list(
        "id_puntoventa_id", "id_producto_id"
    ).distinct():
        cc = _normalize_str(cc)
        mat = _normalize_str(mat)
        if mat in prod_map:
            candidates.add((cc, mat))

    for cc, mat in VentaAbastecimiento.objects.values_list(
        "id_puntoventa_id", "id_producto_id"
    ).distinct():
        cc = _normalize_str(cc)
        mat = _normalize_str(mat)
        if mat in prod_map:
            candidates.add((cc, mat))

    today = date.today()
    current_month_start = today.replace(day=1)
    next_month_start = (today + relativedelta(months=1)).replace(day=1)
    months = [(today - relativedelta(months=i)).replace(day=1) for i in range(1, 4)]

    ventas_monthly_dict = {}
    ventas_current_dict = {}
    ventas_all_dict = {}

    ventas_monthly_qs = (
        VentaAbastecimiento.objects.filter(
            fecha_venta__gte=(today - relativedelta(months=3)).replace(day=1),
            fecha_venta__lt=current_month_start,
        )
        .annotate(year=ExtractYear("fecha_venta"), month=ExtractMonth("fecha_venta"))
        .values("id_puntoventa_id", "id_producto_id", "year", "month")
        .annotate(total=Sum("cantidad_venta"))
    )
    for row in ventas_monthly_qs:
        key = (
            _normalize_str(row["id_puntoventa_id"]),
            _normalize_str(row["id_producto_id"]),
            row["year"],
            row["month"],
        )
        ventas_monthly_dict[key] = int(row["total"] or 0)

    ventas_current_qs = (
        VentaAbastecimiento.objects.filter(
            fecha_venta__gte=current_month_start, fecha_venta__lt=next_month_start
        )
        .values("id_puntoventa_id", "id_producto_id")
        .annotate(total=Sum("cantidad_venta"))
    )
    for row in ventas_current_qs:
        key = (
            _normalize_str(row["id_puntoventa_id"]),
            _normalize_str(row["id_producto_id"]),
        )
        ventas_current_dict[key] = int(row["total"] or 0)

    ventas_lists = {}
    for row in VentaAbastecimiento.objects.values(
        "id_puntoventa_id", "id_producto_id", "cantidad_venta"
    ):
        key = (
            _normalize_str(row["id_puntoventa_id"]),
            _normalize_str(row["id_producto_id"]),
        )
        ventas_lists.setdefault(key, []).append(int(row["cantidad_venta"] or 0))
    for key, values in ventas_lists.items():
        if values:
            ventas_all_dict[key] = float(median(values))

    inv_dict = {}
    inv_qs = (
        InventarioAbastecimiento.objects.values("id_puntoventa_id", "id_producto_id")
        .annotate(total=Sum("cantidad_inventario"))
    )
    for row in inv_qs:
        key = (
            _normalize_str(row["id_puntoventa_id"]),
            _normalize_str(row["id_producto_id"]),
        )
        inv_dict[key] = int(row["total"] or 0)

    tra_dict = {}
    tra_qs = (
        TransitoAbastecimiento.objects.values("id_puntoventa_id", "id_producto_id")
        .annotate(total=Sum("cantidad_transito"))
    )
    for row in tra_qs:
        key = (
            _normalize_str(row["id_puntoventa_id"]),
            _normalize_str(row["id_producto_id"]),
        )
        tra_dict[key] = int(row["total"] or 0)

    def candidate_matches_filters(cc, mat):
        if id_puntoventa_filter and cc.lower() not in id_puntoventa_filter:
            return False
        punto = puntos_map.get(cc, {}).get("Punto de Venta", "").lower()
        if punto_venta_filter and punto not in punto_venta_filter:
            return False
        if id_producto_filter and mat.lower() not in id_producto_filter:
            return False
        prod_name = prod_map.get(mat, {}).get("Producto", "").lower()
        if nombre_producto_filter:
            found = False
            for pat in nombre_producto_filter:
                if pat in prod_name:
                    found = True
                    break
            if not found:
                return False
        marca = prod_map.get(mat, {}).get("Marca", "").lower()
        if marcas_filter and marca not in marcas_filter:
            return False
        canal_val = puntos_map.get(cc, {}).get("Canal o Regional", "").lower()
        if canal_regional_filter and canal_val not in canal_regional_filter:
            return False
        return True

    records = []
    for cc, mat in sorted(candidates):
        if not candidate_matches_filters(cc, mat):
            continue

        month_totals = []
        for d in months:
            key = (cc, mat, d.year, d.month)
            month_totals.append(int(ventas_monthly_dict.get(key, 0)))

        ventas_pasado = month_totals[0]
        ventas_promedio = round(sum(month_totals) / 3.0, 2)
        ventas_actual = int(ventas_current_dict.get((cc, mat), 0))

        inventario = int(inv_dict.get((cc, mat), 0))
        transitos = int(tra_dict.get((cc, mat), 0))

        mediana = ventas_all_dict.get((cc, mat))
        if mediana is not None:
            mediana = float(round(mediana, 2))

        if not (
            inventario > 0
            or transitos > 0
            or ventas_actual > 0
            or ventas_pasado > 0
            or ventas_promedio > 0
        ):
            continue

        envio_3_meses = round(ventas_promedio - inventario, 2)
        envio_pasadas = round(ventas_pasado - inventario, 2)

        indicador_3_meses = None
        if ventas_promedio != 0:
            indicador_3_meses = round(inventario / ventas_promedio, 4)
        indicador_mes_pasado = None
        if ventas_pasado != 0:
            indicador_mes_pasado = round(inventario / ventas_pasado, 4)

        producto_name = prod_map.get(mat, {}).get("Producto", "")
        marca_name = prod_map.get(mat, {}).get("Marca", "")
        punto_name = puntos_map.get(cc, {}).get("Punto de Venta", "")
        canal_name = puntos_map.get(cc, {}).get("Canal o Regional", "")

        records.append(
            {
                "id_puntoventa": cc,
                "id_producto": mat,
                "nombre_producto": producto_name,
                "marca": marca_name,
                "punto_venta": punto_name,
                "canal_regional": canal_name,
                "Ventas_Mes_Actual": ventas_actual,
                "Ventas_Mes_Pasado": ventas_pasado,
                "Ventas_Promedio_3_Meses": ventas_promedio,
                "Mediana": mediana,
                "Inventario": inventario,
                "Transitos": transitos,
                "Indicador_3_Meses": indicador_3_meses,
                "Indicador_Mes_Pasado": indicador_mes_pasado,
                "Envio_3_Meses": envio_3_meses,
                "Envio_Pasadas": envio_pasadas,
            }
        )

    records = sorted(
        records,
        key=lambda x: (
            -(x.get("Envio_Pasadas") or 0),
            x.get("id_puntoventa") or "",
            x.get("id_producto") or "",
        ),
    )

    total = len(records)
    total_pages = (total + page_size - 1) // page_size if page_size > 0 else 1
    if page > total_pages and total_pages > 0:
        page = total_pages
    start = (page - 1) * page_size
    end = start + page_size
    page_records = records[start:end]

    return JsonResponse(
        {
            "records": page_records,
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": total_pages,
        }
    )
