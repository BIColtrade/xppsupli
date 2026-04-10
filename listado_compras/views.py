from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher
from io import BytesIO

from django.db import IntegrityError, transaction
from django.db.models import Count
from django.http import HttpResponse
from django.shortcuts import redirect, render
from openpyxl import Workbook, load_workbook

from user.jwt_utils import get_user_from_request

from .models import listado_productos_internacionales as ListadoProductosInternacionales
from .models import listado_productos_nacionales as ListadoProductosNacionales
from .models import listado_productos_supli as ListadoProductosSupli

LISTADO_COMPRAS_GROUP = "listadocompras"


def _require_listado_compras_group(request):
    user = get_user_from_request(request)
    if user is None:
        return None, redirect("login")
    if not user.groups.filter(name=LISTADO_COMPRAS_GROUP).exists():
        return user, render(request, "acceso_no_permitido.html", status=403)
    return user, None


def _build_xlsx_response(queryset, filename, sheet_name):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    fields = list(queryset.model._meta.fields)
    headers = [field.name for field in fields]
    ws.append(headers)
    for obj in queryset:
        ws.append([field.value_from_object(obj) for field in fields])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f"attachment; filename={filename}"
    return response


def _build_xlsx_from_rows(headers, rows, filename, sheet_name):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append(row)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f"attachment; filename={filename}"
    return response


def _load_xlsx_rows(archivo):
    wb = load_workbook(archivo, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("El archivo esta vacio.")
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    return headers, rows[1:]


def _row_is_empty(row):
    return not row or all(cell is None or str(cell).strip() == "" for cell in row)


def _get_cell(row, idx):
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _similarity_percent(left, right):
    left_text = "" if left is None else str(left).strip().lower()
    right_text = "" if right is None else str(right).strip().lower()
    if not left_text or not right_text:
        return None
    return round(SequenceMatcher(None, left_text, right_text).ratio() * 100, 2)


def _build_cruce_internacional_context(selected_upc):
    selected_info = ListadoProductosSupli.objects.filter(UPC=selected_upc).first()
    selected_count = ListadoProductosInternacionales.objects.filter(upc=selected_upc).count()

    detalles = []
    detalles_qs = ListadoProductosInternacionales.objects.filter(upc=selected_upc).order_by(
        "costo_con_factor_logistico",
        "costo",
    )
    detalles_list = list(detalles_qs)
    costos = [producto.costo for producto in detalles_list]
    cantidades = [producto.cantidad_disponible for producto in detalles_list]
    min_costo = min(costos) if costos else None
    max_cantidad = max(cantidades) if cantidades else None
    best_row = None
    for producto in detalles_list:
        if best_row is None or producto.costo_con_factor_logistico < best_row.costo_con_factor_logistico:
            best_row = producto
    for idx, producto in enumerate(detalles_list, start=1):
        detalles.append(
            {
                "fecha_lista": producto.fecha_lista,
                "nombre": producto.nombre,
                "nombre_supli": selected_info.nombre_producto if selected_info else "",
                "proveedores": producto.proveedores,
                "costo": producto.costo,
                "cantidad_disponible": producto.cantidad_disponible,
                "factor_logistico": producto.factor_logistico,
                "costo_con_factor_logistico": producto.costo_con_factor_logistico,
                "top": idx,
                "es_menor_costo": min_costo is not None and producto.costo == min_costo,
                "es_mayor_cantidad": max_cantidad is not None
                and producto.cantidad_disponible == max_cantidad,
                "similitud_nombre": _similarity_percent(
                    producto.nombre,
                    selected_info.nombre_producto if selected_info else "",
                ),
            }
        )

    mejor_proveedor = None
    if best_row is not None:
        mejor_proveedor = {
            "proveedores": best_row.proveedores,
            "nombre": best_row.nombre,
            "costo_con_factor_logistico": best_row.costo_con_factor_logistico,
            "cantidad_disponible": best_row.cantidad_disponible,
        }

    return {
        "selected_upc": selected_upc,
        "selected_info": selected_info,
        "selected_count": selected_count,
        "detalles": detalles,
        "mejor_proveedor": mejor_proveedor,
    }


def _build_cruce_internacional_general_rows():
    registros = ListadoProductosInternacionales.objects.order_by(
        "upc",
        "costo_con_factor_logistico",
        "costo",
        "id",
    )
    best_by_upc = {}
    for registro in registros:
        if registro.upc not in best_by_upc:
            best_by_upc[registro.upc] = registro

    upc_list = list(best_by_upc.keys())
    supli_map = (
        ListadoProductosSupli.objects.in_bulk(upc_list, field_name="UPC")
        if upc_list
        else {}
    )

    rows = []
    for upc in sorted(best_by_upc.keys()):
        registro = best_by_upc[upc]
        supli = supli_map.get(upc)
        nombre_supli = supli.nombre_producto if supli else ""
        marca_supli = supli.marca_producto if supli else ""
        rows.append(
            {
                "upc": upc,
                "nombre_supli": nombre_supli,
                "nombre_proveedor": registro.nombre,
                "proveedor": registro.proveedores or "",
                "marca_supli": marca_supli,
                "costo": registro.costo,
                "factor_logistico": registro.factor_logistico,
                "costo_con_factor_logistico": registro.costo_con_factor_logistico,
                "cantidad_disponible": registro.cantidad_disponible,
                "similitud_nombre": _similarity_percent(registro.nombre, nombre_supli),
            }
        )

    return rows


def _build_cruce_nacional_context(selected_upc):
    selected_info = ListadoProductosSupli.objects.filter(UPC=selected_upc).first()
    selected_count = ListadoProductosNacionales.objects.filter(upc=selected_upc).count()

    detalles = []
    detalles_qs = ListadoProductosNacionales.objects.filter(upc=selected_upc).order_by(
        "total_costo",
        "costo",
        "id",
    )
    detalles_list = list(detalles_qs)
    totales = [producto.total_costo for producto in detalles_list]
    cantidades = [producto.cantidad_disponible for producto in detalles_list]
    min_total = min(totales) if totales else None
    max_cantidad = max(cantidades) if cantidades else None
    best_row = None
    for producto in detalles_list:
        if best_row is None or producto.total_costo < best_row.total_costo:
            best_row = producto
    for idx, producto in enumerate(detalles_list, start=1):
        detalles.append(
            {
                "nombre": producto.nombre,
                "nombre_supli": selected_info.nombre_producto if selected_info else "",
                "proveedor": producto.proveedor,
                "costo": producto.costo,
                "costos_adicionales": producto.costos_adicionales,
                "total_costo": producto.total_costo,
                "cantidad_disponible": producto.cantidad_disponible,
                "top": idx,
                "es_menor_costo": min_total is not None and producto.total_costo == min_total,
                "es_mayor_cantidad": max_cantidad is not None
                and producto.cantidad_disponible == max_cantidad,
                "similitud_nombre": _similarity_percent(
                    producto.nombre,
                    selected_info.nombre_producto if selected_info else "",
                ),
            }
        )

    mejor_proveedor = None
    if best_row is not None:
        mejor_proveedor = {
            "proveedor": best_row.proveedor,
            "nombre": best_row.nombre,
            "total_costo": best_row.total_costo,
            "cantidad_disponible": best_row.cantidad_disponible,
        }

    return {
        "selected_upc": selected_upc,
        "selected_info": selected_info,
        "selected_count": selected_count,
        "detalles": detalles,
        "mejor_proveedor": mejor_proveedor,
    }


def _build_cruce_nacional_general_rows():
    registros = ListadoProductosNacionales.objects.order_by(
        "upc",
        "total_costo",
        "costo",
        "id",
    )
    best_by_upc = {}
    for registro in registros:
        if registro.upc not in best_by_upc:
            best_by_upc[registro.upc] = registro

    upc_list = list(best_by_upc.keys())
    supli_map = (
        ListadoProductosSupli.objects.in_bulk(upc_list, field_name="UPC")
        if upc_list
        else {}
    )

    rows = []
    for upc in sorted(best_by_upc.keys()):
        registro = best_by_upc[upc]
        supli = supli_map.get(upc)
        nombre_supli = supli.nombre_producto if supli else ""
        marca_supli = supli.marca_producto if supli else ""
        rows.append(
            {
                "upc": upc,
                "nombre_supli": nombre_supli,
                "nombre_proveedor": registro.nombre,
                "proveedor": registro.proveedor or "",
                "marca_supli": marca_supli,
                "costo": registro.costo,
                "costos_adicionales": registro.costos_adicionales,
                "total_costo": registro.total_costo,
                "cantidad_disponible": registro.cantidad_disponible,
                "similitud_nombre": _similarity_percent(registro.nombre, nombre_supli),
            }
        )

    return rows


def home_listado_compras(request):
    user, response = _require_listado_compras_group(request)
    if response is not None:
        return response
    return render(request, "home_listado_compras.html")


def cruce_producto_internacional(request):
    user, response = _require_listado_compras_group(request)
    if response is not None:
        return response

    upc_rows = (
        ListadoProductosInternacionales.objects.values("upc")
        .annotate(total=Count("id"))
        .order_by("upc")
    )
    upc_list = [row["upc"] for row in upc_rows]
    supli_map = (
        ListadoProductosSupli.objects.in_bulk(upc_list, field_name="UPC")
        if upc_list
        else {}
    )

    resumen = []
    for row in upc_rows:
        upc = row["upc"]
        supli = supli_map.get(upc)
        resumen.append(
            {
                "upc": upc,
                "total": row["total"],
                "nombre_producto": supli.nombre_producto if supli else "",
                "marca_producto": supli.marca_producto if supli else "",
            }
        )

    context = {
        "upc_resumen": resumen,
    }

    return render(request, "cruce_producto_internacional.html", context)


def cruce_producto_internacional_detalle(request):
    user, response = _require_listado_compras_group(request)
    if response is not None:
        return response

    selected_upc = request.GET.get("upc", "").strip()
    if not selected_upc:
        return redirect("listado_compras:cruce_producto_internacional")

    context = _build_cruce_internacional_context(selected_upc)

    return render(request, "cruce_producto_internacional_detalle.html", context)


def cruce_producto_internacional_export(request):
    user, response = _require_listado_compras_group(request)
    if response is not None:
        return response

    selected_upc = request.GET.get("upc", "").strip()
    if not selected_upc:
        return redirect("listado_compras:cruce_producto_internacional")

    context = _build_cruce_internacional_context(selected_upc)
    headers = [
        "top",
        "fecha_lista",
        "nombre_producto",
        "nombre_producto_supli",
        "proveedores",
        "costo",
        "cantidad_disponible",
        "factor_logistico",
        "costo_con_factor_logistico",
        "similitud_nombre",
    ]
    rows = []
    for fila in context["detalles"]:
        rows.append(
            [
                fila["top"],
                fila["fecha_lista"],
                fila["nombre"],
                fila["nombre_supli"],
                fila["proveedores"],
                fila["costo"],
                fila["cantidad_disponible"],
                fila["factor_logistico"],
                fila["costo_con_factor_logistico"],
                fila["similitud_nombre"],
            ]
        )

    filename = f"cruce_internacional_{selected_upc}.xlsx"
    return _build_xlsx_from_rows(headers, rows, filename, "CruceInternacional")


def cruce_producto_internacional_detalle_general(request):
    user, response = _require_listado_compras_group(request)
    if response is not None:
        return response

    rows = _build_cruce_internacional_general_rows()
    context = {"resumen": rows}
    return render(request, "cruce_producto_internacional_detalle_General.html", context)


def cruce_producto_internacional_detalle_general_export(request):
    user, response = _require_listado_compras_group(request)
    if response is not None:
        return response

    rows = _build_cruce_internacional_general_rows()
    headers = [
        "upc",
        "nombre_supli",
        "nombre_producto_proveedor",
        "proveedor",
        "marca_supli",
        "costo",
        "factor_logistico",
        "costo_con_factor_logistico",
        "cantidad_disponible",
        "similitud_nombre",
    ]
    data = []
    for fila in rows:
        data.append(
            [
                fila["upc"],
                fila["nombre_supli"],
                fila["nombre_proveedor"],
                fila["proveedor"],
                fila["marca_supli"],
                fila["costo"],
                fila["factor_logistico"],
                fila["costo_con_factor_logistico"],
                fila["cantidad_disponible"],
                fila["similitud_nombre"],
            ]
        )

    return _build_xlsx_from_rows(
        headers, data, "cruce_internacional_general.xlsx", "CruceGeneral"
    )


def cruce_producto_nacional(request):
    user, response = _require_listado_compras_group(request)
    if response is not None:
        return response

    upc_rows = (
        ListadoProductosNacionales.objects.values("upc")
        .annotate(total=Count("id"))
        .order_by("upc")
    )
    upc_list = [row["upc"] for row in upc_rows]
    supli_map = (
        ListadoProductosSupli.objects.in_bulk(upc_list, field_name="UPC")
        if upc_list
        else {}
    )

    resumen = []
    for row in upc_rows:
        upc = row["upc"]
        supli = supli_map.get(upc)
        resumen.append(
            {
                "upc": upc,
                "total": row["total"],
                "nombre_producto": supli.nombre_producto if supli else "",
                "marca_producto": supli.marca_producto if supli else "",
            }
        )

    context = {
        "upc_resumen": resumen,
    }

    return render(request, "cruce_producto_nacional.html", context)


def cruce_producto_nacional_detalle(request):
    user, response = _require_listado_compras_group(request)
    if response is not None:
        return response

    selected_upc = request.GET.get("upc", "").strip()
    if not selected_upc:
        return redirect("listado_compras:cruce_producto_nacional")

    context = _build_cruce_nacional_context(selected_upc)

    return render(request, "cruce_producto_nacional_detalle.html", context)


def cruce_producto_nacional_export(request):
    user, response = _require_listado_compras_group(request)
    if response is not None:
        return response

    selected_upc = request.GET.get("upc", "").strip()
    if not selected_upc:
        return redirect("listado_compras:cruce_producto_nacional")

    context = _build_cruce_nacional_context(selected_upc)
    headers = [
        "top",
        "nombre_producto",
        "nombre_producto_supli",
        "proveedor",
        "costo",
        "costos_adicionales",
        "total_costo",
        "cantidad_disponible",
        "similitud_nombre",
    ]
    rows = []
    for fila in context["detalles"]:
        rows.append(
            [
                fila["top"],
                fila["nombre"],
                fila["nombre_supli"],
                fila["proveedor"],
                fila["costo"],
                fila["costos_adicionales"],
                fila["total_costo"],
                fila["cantidad_disponible"],
                fila["similitud_nombre"],
            ]
        )

    filename = f"cruce_nacional_{selected_upc}.xlsx"
    return _build_xlsx_from_rows(headers, rows, filename, "CruceNacional")


def cruce_producto_nacional_detalle_general(request):
    user, response = _require_listado_compras_group(request)
    if response is not None:
        return response

    rows = _build_cruce_nacional_general_rows()
    context = {"resumen": rows}
    return render(request, "cruce_producto_nacional_detalle_General.html", context)


def cruce_producto_nacional_detalle_general_export(request):
    user, response = _require_listado_compras_group(request)
    if response is not None:
        return response

    rows = _build_cruce_nacional_general_rows()
    headers = [
        "upc",
        "nombre_supli",
        "nombre_producto_proveedor",
        "proveedor",
        "marca_supli",
        "costo",
        "costos_adicionales",
        "total_costo",
        "cantidad_disponible",
        "similitud_nombre",
    ]
    data = []
    for fila in rows:
        data.append(
            [
                fila["upc"],
                fila["nombre_supli"],
                fila["nombre_proveedor"],
                fila["proveedor"],
                fila["marca_supli"],
                fila["costo"],
                fila["costos_adicionales"],
                fila["total_costo"],
                fila["cantidad_disponible"],
                fila["similitud_nombre"],
            ]
        )

    return _build_xlsx_from_rows(
        headers, data, "cruce_nacional_general.xlsx", "CruceNacionalGeneral"
    )


def listado_productos_supli(request):
    user, response = _require_listado_compras_group(request)
    if response is not None:
        return response
    return render(request, "listado_productos_supli.html")


def crud_productos_supli(request):
    user, response = _require_listado_compras_group(request)
    if response is not None:
        return response
    context = {"productos": ListadoProductosSupli.objects.order_by("nombre_producto")}
    flash_success = request.session.pop("productos_supli_success", None)
    if flash_success:
        context["success"] = flash_success
    if request.method == "POST":
        action = request.POST.get("action", "").strip()
        if action == "create":
            upc = request.POST.get("UPC", "").strip()
            nombre_producto = request.POST.get("nombre_producto", "").strip()
            marca_producto = request.POST.get("marca_producto", "").strip()
            if not upc:
                context["error"] = "El UPC es obligatorio."
            elif not nombre_producto:
                context["error"] = "El nombre del producto es obligatorio."
            elif not marca_producto:
                context["error"] = "La marca del producto es obligatoria."
            else:
                try:
                    ListadoProductosSupli.objects.create(
                        UPC=upc,
                        nombre_producto=nombre_producto,
                        marca_producto=marca_producto,
                    )
                except IntegrityError:
                    context["error"] = "Ya existe un producto con ese UPC."
                else:
                    request.session["productos_supli_success"] = "Producto creado."
                    return redirect("listado_compras:crud_productos_supli")
        else:
            producto_upc = request.POST.get("producto_upc", "").strip()
            if not producto_upc:
                context["error"] = "Producto no valido."
            else:
                try:
                    producto = ListadoProductosSupli.objects.get(pk=producto_upc)
                except ListadoProductosSupli.DoesNotExist:
                    context["error"] = "Producto no encontrado."
                else:
                    if action == "delete":
                        try:
                            producto.delete()
                            request.session["productos_supli_success"] = "Producto eliminado."
                            return redirect("listado_compras:crud_productos_supli")
                        except Exception:
                            context["error"] = "No se pudo eliminar el producto."
                    elif action == "save":
                        nombre_producto = request.POST.get("nombre_producto", "").strip()
                        marca_producto = request.POST.get("marca_producto", "").strip()
                        if not nombre_producto:
                            context["error"] = "El nombre del producto es obligatorio."
                        elif not marca_producto:
                            context["error"] = "La marca del producto es obligatoria."
                        else:
                            producto.nombre_producto = nombre_producto
                            producto.marca_producto = marca_producto
                            producto.save()
                            request.session["productos_supli_success"] = "Producto actualizado."
                            return redirect("listado_compras:crud_productos_supli")

        context["productos"] = ListadoProductosSupli.objects.order_by("nombre_producto")

    return render(request, "crud_productos_supli.html", context)


def crud_listado_internacional(request):
    user, response = _require_listado_compras_group(request)
    if response is not None:
        return response
    context = {"productos": ListadoProductosInternacionales.objects.order_by("nombre")}
    flash_success = request.session.pop("productos_internacionales_success", None)
    if flash_success:
        context["success"] = flash_success
    if request.method == "POST":
        action = request.POST.get("action", "").strip()
        if action == "create":
            fecha_lista = request.POST.get("fecha_lista", "").strip()
            upc = request.POST.get("upc", "").strip()
            nombre = request.POST.get("nombre", "").strip()
            costo_raw = request.POST.get("costo", "").strip()
            cantidad_raw = request.POST.get("cantidad_disponible", "").strip()
            proveedores = request.POST.get("proveedores", "").strip()
            factor_raw = request.POST.get("factor_logistico", "").strip()
            if not fecha_lista:
                context["error"] = "La fecha de lista es obligatoria."
            elif not upc:
                context["error"] = "El UPC es obligatorio."
            elif not nombre:
                context["error"] = "El nombre es obligatorio."
            elif not costo_raw:
                context["error"] = "El costo es obligatorio."
            elif not cantidad_raw:
                context["error"] = "La cantidad disponible es obligatoria."
            elif not proveedores:
                context["error"] = "Los proveedores son obligatorios."
            elif factor_raw == "":
                context["error"] = "El factor logistico es obligatorio."
            else:
                try:
                    costo = Decimal(costo_raw)
                    cantidad = int(cantidad_raw)
                    factor = Decimal(factor_raw)
                except (InvalidOperation, ValueError):
                    context["error"] = "Costo, cantidad o factor logistico no son validos."
                else:
                    try:
                        ListadoProductosInternacionales.objects.create(
                            fecha_lista=fecha_lista,
                            upc=upc,
                            nombre=nombre,
                            costo=costo,
                            cantidad_disponible=cantidad,
                            proveedores=proveedores,
                            factor_logistico=factor,
                        )
                    except IntegrityError:
                        context["error"] = "No se pudo crear el producto."
                    else:
                        request.session["productos_internacionales_success"] = (
                            "Producto creado."
                        )
                        return redirect("listado_compras:crud_listado_internacional")
        else:
            producto_id = request.POST.get("producto_id", "").strip()
            if not producto_id:
                context["error"] = "Producto no valido."
            else:
                try:
                    producto = ListadoProductosInternacionales.objects.get(pk=producto_id)
                except ListadoProductosInternacionales.DoesNotExist:
                    context["error"] = "Producto no encontrado."
                else:
                    if action == "delete":
                        try:
                            producto.delete()
                            request.session["productos_internacionales_success"] = "Producto eliminado."
                            return redirect("listado_compras:crud_listado_internacional")
                        except Exception:
                            context["error"] = "No se pudo eliminar el producto."
                    elif action == "save":
                        fecha_lista = request.POST.get("fecha_lista", "").strip()
                        nombre = request.POST.get("nombre", "").strip()
                        costo_raw = request.POST.get("costo", "").strip()
                        cantidad_raw = request.POST.get("cantidad_disponible", "").strip()
                        proveedores = request.POST.get("proveedores", "").strip()
                        factor_raw = request.POST.get("factor_logistico", "").strip()
                        if not fecha_lista:
                            context["error"] = "La fecha de lista es obligatoria."
                        elif not nombre:
                            context["error"] = "El nombre es obligatorio."
                        elif not costo_raw:
                            context["error"] = "El costo es obligatorio."
                        elif not cantidad_raw:
                            context["error"] = "La cantidad disponible es obligatoria."
                        elif not proveedores:
                            context["error"] = "Los proveedores son obligatorios."
                        elif factor_raw == "":
                            context["error"] = "El factor logistico es obligatorio."
                        else:
                            try:
                                costo = Decimal(costo_raw)
                                cantidad = int(cantidad_raw)
                                factor = Decimal(factor_raw)
                            except (InvalidOperation, ValueError):
                                context["error"] = "Costo, cantidad o factor logistico no son validos."
                            else:
                                producto.fecha_lista = fecha_lista
                                producto.nombre = nombre
                                producto.costo = costo
                                producto.cantidad_disponible = cantidad
                                producto.proveedores = proveedores
                                producto.factor_logistico = factor
                                producto.save()
                                request.session["productos_internacionales_success"] = "Producto actualizado."
                                return redirect("listado_compras:crud_listado_internacional")

        context["productos"] = ListadoProductosInternacionales.objects.order_by("nombre")

    return render(request, "crud_listado_internacional.html", context)


def crud_listado_nacional(request):
    user, response = _require_listado_compras_group(request)
    if response is not None:
        return response
    context = {"productos": ListadoProductosNacionales.objects.order_by("nombre")}
    flash_success = request.session.pop("productos_nacionales_success", None)
    if flash_success:
        context["success"] = flash_success
    if request.method == "POST":
        action = request.POST.get("action", "").strip()
        if action == "create":
            upc = request.POST.get("upc", "").strip()
            nombre = request.POST.get("nombre", "").strip()
            costo_raw = request.POST.get("costo", "").strip()
            cantidad_raw = request.POST.get("cantidad_disponible", "").strip()
            proveedor = request.POST.get("proveedor", "").strip()
            adicionales_raw = request.POST.get("costos_adicionales", "").strip()
            if not upc:
                context["error"] = "El UPC es obligatorio."
            elif not nombre:
                context["error"] = "El nombre es obligatorio."
            elif not costo_raw:
                context["error"] = "El costo es obligatorio."
            elif not cantidad_raw:
                context["error"] = "La cantidad disponible es obligatoria."
            elif not proveedor:
                context["error"] = "El proveedor es obligatorio."
            elif adicionales_raw == "":
                context["error"] = "Los costos adicionales son obligatorios."
            else:
                try:
                    costo = _parse_decimal(costo_raw)
                    adicionales = _parse_decimal(adicionales_raw)
                    cantidad = int(cantidad_raw)
                except (InvalidOperation, ValueError, TypeError):
                    context["error"] = "Costo, cantidad o costos adicionales no son validos."
                else:
                    if costo is None or adicionales is None:
                        context["error"] = "El costo y los costos adicionales son obligatorios."
                    else:
                        try:
                            ListadoProductosNacionales.objects.create(
                                upc=upc,
                                nombre=nombre,
                                costo=costo,
                                cantidad_disponible=cantidad,
                                proveedor=proveedor,
                                costos_adicionales=adicionales,
                            )
                        except IntegrityError:
                            context["error"] = "Ya existe un producto con ese UPC."
                        else:
                            request.session["productos_nacionales_success"] = "Producto creado."
                            return redirect("listado_compras:crud_listado_nacional")
        else:
            producto_id = request.POST.get("producto_id", "").strip()
            if not producto_id:
                context["error"] = "Producto no valido."
            else:
                try:
                    producto = ListadoProductosNacionales.objects.get(pk=producto_id)
                except ListadoProductosNacionales.DoesNotExist:
                    context["error"] = "Producto no encontrado."
                else:
                    if action == "delete":
                        try:
                            producto.delete()
                            request.session["productos_nacionales_success"] = "Producto eliminado."
                            return redirect("listado_compras:crud_listado_nacional")
                        except Exception:
                            context["error"] = "No se pudo eliminar el producto."
                    elif action == "save":
                        nombre = request.POST.get("nombre", "").strip()
                        costo_raw = request.POST.get("costo", "").strip()
                        cantidad_raw = request.POST.get("cantidad_disponible", "").strip()
                        proveedor = request.POST.get("proveedor", "").strip()
                        adicionales_raw = request.POST.get("costos_adicionales", "").strip()
                        if not nombre:
                            context["error"] = "El nombre es obligatorio."
                        elif not costo_raw:
                            context["error"] = "El costo es obligatorio."
                        elif not cantidad_raw:
                            context["error"] = "La cantidad disponible es obligatoria."
                        elif not proveedor:
                            context["error"] = "El proveedor es obligatorio."
                        elif adicionales_raw == "":
                            context["error"] = "Los costos adicionales son obligatorios."
                        else:
                            try:
                                costo = _parse_decimal(costo_raw)
                                adicionales = _parse_decimal(adicionales_raw)
                                cantidad = int(cantidad_raw)
                            except (InvalidOperation, ValueError, TypeError):
                                context["error"] = "Costo, cantidad o costos adicionales no son validos."
                            else:
                                if costo is None or adicionales is None:
                                    context["error"] = "El costo y los costos adicionales son obligatorios."
                                else:
                                    producto.nombre = nombre
                                    producto.costo = costo
                                    producto.cantidad_disponible = cantidad
                                    producto.proveedor = proveedor
                                    producto.costos_adicionales = adicionales
                                    producto.save()
                                    request.session["productos_nacionales_success"] = (
                                        "Producto actualizado."
                                    )
                                    return redirect("listado_compras:crud_listado_nacional")

        context["productos"] = ListadoProductosNacionales.objects.order_by("nombre")

    return render(request, "crud_listado_nacional.html", context)


def productos_supli_export(request):
    user, response = _require_listado_compras_group(request)
    if response is not None:
        return response
    queryset = ListadoProductosSupli.objects.order_by("nombre_producto")
    return _build_xlsx_response(queryset, "productos_supli.xlsx", "ProductosSupli")


def productos_supli_import(request):
    user, response = _require_listado_compras_group(request)
    if response is not None:
        return response
    if request.method != "POST":
        return redirect("listado_compras:crud_productos_supli")
    archivo = request.FILES.get("archivo")
    if not archivo:
        return redirect("listado_compras:crud_productos_supli")
    try:
        headers, rows = _load_xlsx_rows(archivo)
        required = ["UPC", "nombre_producto", "marca_producto"]
        for name in required:
            if name not in headers:
                raise ValueError(f"Falta la columna {name}.")
        header_index = {name: idx for idx, name in enumerate(headers)}

        existing = ListadoProductosSupli.objects.in_bulk()
        productos_a_crear = []
        productos_a_actualizar = []
        seen_upc = set()

        with transaction.atomic():
            for row in rows:
                if _row_is_empty(row):
                    continue
                upc = _get_cell(row, header_index["UPC"])
                nombre_producto = _get_cell(row, header_index["nombre_producto"])
                marca_producto = _get_cell(row, header_index["marca_producto"])

                upc = "" if upc is None else str(upc).strip()
                nombre_producto = "" if nombre_producto is None else str(nombre_producto).strip()
                marca_producto = "" if marca_producto is None else str(marca_producto).strip()

                if not upc or not nombre_producto or not marca_producto:
                    raise ValueError("Faltan campos requeridos en el archivo.")
                if upc in seen_upc:
                    raise ValueError(f"UPC duplicado en el archivo: {upc}")
                seen_upc.add(upc)

                if upc in existing:
                    producto = existing[upc]
                    producto.nombre_producto = nombre_producto
                    producto.marca_producto = marca_producto
                    productos_a_actualizar.append(producto)
                else:
                    productos_a_crear.append(
                        ListadoProductosSupli(
                            UPC=upc,
                            nombre_producto=nombre_producto,
                            marca_producto=marca_producto,
                        )
                    )

            if productos_a_crear:
                ListadoProductosSupli.objects.bulk_create(productos_a_crear, batch_size=500)
            if productos_a_actualizar:
                ListadoProductosSupli.objects.bulk_update(
                    productos_a_actualizar,
                    ["nombre_producto", "marca_producto"],
                    batch_size=500,
                )

        request.session["productos_supli_success"] = "Importacion completada."
        return redirect("listado_compras:crud_productos_supli")
    except Exception as exc:
        context = {
            "productos": ListadoProductosSupli.objects.order_by("nombre_producto"),
            "error": f"Error al importar: {exc}",
        }
        return render(request, "crud_productos_supli.html", context)


def productos_internacionales_export(request):
    user, response = _require_listado_compras_group(request)
    if response is not None:
        return response
    queryset = ListadoProductosInternacionales.objects.order_by("nombre")
    return _build_xlsx_response(
        queryset, "productos_internacionales.xlsx", "ProductosInternacionales"
    )


def productos_nacionales_export(request):
    user, response = _require_listado_compras_group(request)
    if response is not None:
        return response
    queryset = ListadoProductosNacionales.objects.order_by("nombre")
    return _build_xlsx_response(
        queryset, "productos_nacionales.xlsx", "ProductosNacionales"
    )


def _parse_decimal(value):
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value).replace(",", ".").strip())


def productos_internacionales_import(request):
    user, response = _require_listado_compras_group(request)
    if response is not None:
        return response
    if request.method != "POST":
        return redirect("listado_compras:crud_listado_internacional")
    archivo = request.FILES.get("archivo")
    if not archivo:
        return redirect("listado_compras:crud_listado_internacional")
    try:
        headers, rows = _load_xlsx_rows(archivo)
        required = [
            "fecha_lista",
            "upc",
            "nombre",
            "costo",
            "cantidad_disponible",
            "proveedores",
            "factor_logistico",
        ]
        for name in required:
            if name not in headers:
                raise ValueError(f"Falta la columna {name}.")
        header_index = {name: idx for idx, name in enumerate(headers)}

        productos_a_crear = []

        with transaction.atomic():
            for row in rows:
                if _row_is_empty(row):
                    continue
                fecha_lista = _get_cell(row, header_index["fecha_lista"])
                upc = _get_cell(row, header_index["upc"])
                nombre = _get_cell(row, header_index["nombre"])
                costo = _get_cell(row, header_index["costo"])
                cantidad = _get_cell(row, header_index["cantidad_disponible"])
                proveedores = _get_cell(row, header_index["proveedores"])
                factor = _get_cell(row, header_index["factor_logistico"])

                upc = "" if upc is None else str(upc).strip()
                nombre = "" if nombre is None else str(nombre).strip()
                proveedores = "" if proveedores is None else str(proveedores).strip()

                if not fecha_lista or not upc or not nombre or not proveedores:
                    raise ValueError("Faltan campos requeridos en el archivo.")
                try:
                    costo_dec = _parse_decimal(costo)
                    factor_dec = _parse_decimal(factor)
                    cantidad_int = int(cantidad)
                except (InvalidOperation, ValueError, TypeError):
                    raise ValueError("Costo, cantidad o factor logistico no son validos.")

                if costo_dec is None or factor_dec is None:
                    raise ValueError("Costo y factor logistico son obligatorios.")

                costo_factor = costo_dec + (costo_dec * factor_dec / Decimal("100.00"))
                productos_a_crear.append(
                    ListadoProductosInternacionales(
                        fecha_lista=fecha_lista,
                        upc=upc,
                        nombre=nombre,
                        costo=costo_dec,
                        cantidad_disponible=cantidad_int,
                        proveedores=proveedores,
                        factor_logistico=factor_dec,
                        costo_con_factor_logistico=costo_factor,
                    )
                )

            if productos_a_crear:
                ListadoProductosInternacionales.objects.bulk_create(
                    productos_a_crear, batch_size=500
                )

        request.session["productos_internacionales_success"] = "Importacion completada."
        return redirect("listado_compras:crud_listado_internacional")
    except Exception as exc:
        context = {
            "productos": ListadoProductosInternacionales.objects.order_by("nombre"),
            "error": f"Error al importar: {exc}",
        }
        return render(request, "crud_listado_internacional.html", context)


def productos_nacionales_import(request):
    user, response = _require_listado_compras_group(request)
    if response is not None:
        return response
    if request.method != "POST":
        return redirect("listado_compras:crud_listado_nacional")
    archivo = request.FILES.get("archivo")
    if not archivo:
        return redirect("listado_compras:crud_listado_nacional")
    try:
        headers, rows = _load_xlsx_rows(archivo)
        required = [
            "upc",
            "nombre",
            "costo",
            "cantidad_disponible",
            "proveedor",
            "costos_adicionales",
        ]
        for name in required:
            if name not in headers:
                raise ValueError(f"Falta la columna {name}.")
        header_index = {name: idx for idx, name in enumerate(headers)}

        existing = ListadoProductosNacionales.objects.in_bulk(field_name="upc")
        productos_a_crear = []
        productos_a_actualizar = []
        seen_upc = set()

        with transaction.atomic():
            for row in rows:
                if _row_is_empty(row):
                    continue
                upc = _get_cell(row, header_index["upc"])
                nombre = _get_cell(row, header_index["nombre"])
                costo = _get_cell(row, header_index["costo"])
                cantidad = _get_cell(row, header_index["cantidad_disponible"])
                proveedor = _get_cell(row, header_index["proveedor"])
                adicionales = _get_cell(row, header_index["costos_adicionales"])

                upc = "" if upc is None else str(upc).strip()
                nombre = "" if nombre is None else str(nombre).strip()
                proveedor = "" if proveedor is None else str(proveedor).strip()

                if not upc or not nombre or not proveedor:
                    raise ValueError("Faltan campos requeridos en el archivo.")
                if upc in seen_upc:
                    raise ValueError(f"UPC duplicado en el archivo: {upc}")
                seen_upc.add(upc)

                try:
                    costo_dec = _parse_decimal(costo)
                    adicionales_dec = _parse_decimal(adicionales)
                    cantidad_int = int(cantidad)
                except (InvalidOperation, ValueError, TypeError):
                    raise ValueError("Costo, cantidad o costos adicionales no son validos.")

                if costo_dec is None or adicionales_dec is None:
                    raise ValueError("Costo y costos adicionales son obligatorios.")

                total_costo = costo_dec + adicionales_dec

                if upc in existing:
                    producto = existing[upc]
                    producto.nombre = nombre
                    producto.costo = costo_dec
                    producto.cantidad_disponible = cantidad_int
                    producto.proveedor = proveedor
                    producto.costos_adicionales = adicionales_dec
                    producto.total_costo = total_costo
                    productos_a_actualizar.append(producto)
                else:
                    productos_a_crear.append(
                        ListadoProductosNacionales(
                            upc=upc,
                            nombre=nombre,
                            costo=costo_dec,
                            cantidad_disponible=cantidad_int,
                            proveedor=proveedor,
                            costos_adicionales=adicionales_dec,
                            total_costo=total_costo,
                        )
                    )

            if productos_a_crear:
                ListadoProductosNacionales.objects.bulk_create(
                    productos_a_crear, batch_size=500
                )
            if productos_a_actualizar:
                ListadoProductosNacionales.objects.bulk_update(
                    productos_a_actualizar,
                    [
                        "nombre",
                        "costo",
                        "cantidad_disponible",
                        "proveedor",
                        "costos_adicionales",
                        "total_costo",
                    ],
                    batch_size=500,
                )

        request.session["productos_nacionales_success"] = "Importacion completada."
        return redirect("listado_compras:crud_listado_nacional")
    except Exception as exc:
        context = {
            "productos": ListadoProductosNacionales.objects.order_by("nombre"),
            "error": f"Error al importar: {exc}",
        }
        return render(request, "crud_listado_nacional.html", context)

