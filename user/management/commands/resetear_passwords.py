"""Rotacion masiva de contrasenas + exporte de credenciales a Excel.

Uso tipico:

    # 1. Simulacion: no toca la base, solo muestra a quien afectaria
    python manage.py resetear_passwords

    # 2. Ejecucion real (pide --confirmar a proposito)
    python manage.py resetear_passwords --confirmar

    # 3. Sin tocar a los superusuarios / staff
    python manage.py resetear_passwords --confirmar --excluir-superusuarios

El Excel queda con la contrasena EN TEXTO PLANO (es la unica forma de
entregarsela a cada persona: en la base solo se guarda el hash). Trata ese
archivo como material sensible: entregalo por canal privado y borralo cuando
todos hayan entrado.
"""

import secrets
import string
from datetime import datetime
from pathlib import Path

from django.conf import settings
from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ValidationError
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from user.models import Usuario


# Alfabeto sin caracteres ambiguos (0/O, 1/l/I) para que nadie se equivoque
# al escribir la clave a mano.
MAYUSCULAS = "ABCDEFGHJKLMNPQRSTUVWXYZ"
MINUSCULAS = "abcdefghijkmnopqrstuvwxyz"
DIGITOS = "23456789"
SIMBOLOS = "!#$%&*+-?@"
ALFABETO = MAYUSCULAS + MINUSCULAS + DIGITOS + SIMBOLOS


def generar_password(longitud=12, usuario=None):
    """Genera una contrasena aleatoria que pasa los validadores de Django.

    Garantiza al menos una mayuscula, una minuscula, un digito y un simbolo.
    """
    longitud = max(longitud, 8)
    for _ in range(200):
        base = [
            secrets.choice(MAYUSCULAS),
            secrets.choice(MINUSCULAS),
            secrets.choice(DIGITOS),
            secrets.choice(SIMBOLOS),
        ]
        base += [secrets.choice(ALFABETO) for _ in range(longitud - len(base))]
        secrets.SystemRandom().shuffle(base)
        candidata = "".join(base)
        try:
            validate_password(candidata, user=usuario)
        except ValidationError:
            continue
        return candidata
    raise CommandError("No fue posible generar una contrasena valida; sube --longitud.")


COLUMNAS = [
    ("ID", lambda u: u.pk),
    ("Nombre", lambda u: u.nombre or ""),
    ("Apellido", lambda u: u.apellido or ""),
    ("Email (usuario de ingreso)", lambda u: u.email),
    ("Username", lambda u: u.username or ""),
    ("Contrasena nueva", None),  # se llena aparte
    ("Area", lambda u: u.get_area_display() if u.area else ""),
    ("Cargo", lambda u: u.cargo or ""),
    ("Tipo de usuario", lambda u: u.get_tipo_usuario_display() if u.tipo_usuario else ""),
    ("Jefe directo", lambda u: str(u.jefe_directo) if u.jefe_directo_id else ""),
    ("Telefono", lambda u: u.telefono or ""),
    ("Edad", lambda u: u.edad if u.edad is not None else ""),
    ("Activo", lambda u: "Si" if u.is_active else "No"),
    ("Staff", lambda u: "Si" if u.is_staff else "No"),
    ("Superusuario", lambda u: "Si" if u.is_superuser else "No"),
    ("Ultimo ingreso", lambda u: u.last_login.strftime("%d/%m/%Y %H:%M") if u.last_login else ""),
]


class Command(BaseCommand):
    help = (
        "Asigna una contrasena nueva y distinta a cada usuario y exporta un Excel "
        "con sus datos y la contrasena generada."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--confirmar", action="store_true",
            help="Ejecuta el cambio real. Sin esta bandera solo simula (dry-run).",
        )
        parser.add_argument(
            "--salida", default=None,
            help="Ruta del .xlsx a generar. Por defecto backups/credenciales_<fecha>.xlsx",
        )
        parser.add_argument(
            "--longitud", type=int, default=12,
            help="Longitud de la contrasena generada (minimo 8, default 12).",
        )
        parser.add_argument(
            "--excluir-superusuarios", action="store_true",
            help="No cambia la contrasena de superusuarios ni de staff.",
        )
        parser.add_argument(
            "--solo-activos", action="store_true",
            help="Solo usuarios con is_active=True.",
        )

    def handle(self, *args, **options):
        confirmar = options["confirmar"]
        longitud = options["longitud"]

        usuarios = Usuario.objects.select_related("jefe_directo").order_by("area", "nombre")
        if options["excluir_superusuarios"]:
            usuarios = usuarios.exclude(is_superuser=True).exclude(is_staff=True)
        if options["solo_activos"]:
            usuarios = usuarios.filter(is_active=True)
        usuarios = list(usuarios)

        if not usuarios:
            raise CommandError("No hay usuarios que coincidan con los filtros.")

        db = settings.DATABASES["default"]
        self.stdout.write(f"Base de datos: {db['ENGINE'].split('.')[-1]} -> {db.get('NAME')}")
        self.stdout.write(f"Usuarios afectados: {len(usuarios)}")

        if not confirmar:
            self.stdout.write(self.style.WARNING(
                "\nSIMULACION (dry-run): no se cambio ninguna contrasena ni se genero Excel."
            ))
            for u in usuarios[:10]:
                self.stdout.write(f"  - {u.email} ({u.get_area_display() if u.area else 'sin area'})")
            if len(usuarios) > 10:
                self.stdout.write(f"  ... y {len(usuarios) - 10} mas")
            self.stdout.write(self.style.WARNING(
                "\nVuelve a correrlo con --confirmar para aplicar el cambio."
            ))
            return

        # ---- Generacion (contrasena distinta para cada persona) ----
        generadas = set()
        credenciales = []
        for u in usuarios:
            while True:
                clave = generar_password(longitud, usuario=u)
                if clave not in generadas:
                    break
            generadas.add(clave)
            credenciales.append((u, clave))

        # ---- Escritura en base (hash, nunca texto plano) ----
        with transaction.atomic():
            for u, clave in credenciales:
                u.set_password(clave)
                u.save(update_fields=["password"])

        ruta = self._exportar_excel(credenciales, options["salida"])

        self.stdout.write(self.style.SUCCESS(
            f"\nListo: {len(credenciales)} contrasenas cambiadas (todas distintas)."
        ))
        self.stdout.write(self.style.SUCCESS(f"Excel generado: {ruta}"))
        self.stdout.write(self.style.WARNING(
            "El archivo contiene contrasenas en texto plano. Entregalo por canal "
            "privado y borralo cuando todos hayan ingresado."
        ))

    # ------------------------------------------------------------------
    def _exportar_excel(self, credenciales, salida):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        if salida:
            ruta = Path(salida)
        else:
            marca = datetime.now().strftime("%Y%m%d_%H%M")
            carpeta = Path(settings.BASE_DIR) / "backups"
            carpeta.mkdir(exist_ok=True)
            ruta = carpeta / f"credenciales_{marca}.xlsx"
        ruta.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Credenciales"

        encabezado_fill = PatternFill("solid", fgColor="19065B")
        encabezado_font = Font(bold=True, color="FFFFFF", size=11)

        titulos = [t for t, _ in COLUMNAS]
        ws.append(titulos)
        for col in range(1, len(titulos) + 1):
            celda = ws.cell(row=1, column=col)
            celda.fill = encabezado_fill
            celda.font = encabezado_font
            celda.alignment = Alignment(horizontal="center", vertical="center")

        col_password = titulos.index("Contrasena nueva") + 1

        for u, clave in credenciales:
            fila = []
            for titulo, getter in COLUMNAS:
                fila.append(clave if getter is None else getter(u))
            ws.append(fila)

        # La contrasena en monoespaciada y como texto, para que Excel no la
        # reinterprete y para que se lea sin confusiones.
        fuente_mono = Font(name="Consolas", bold=True)
        for row in range(2, ws.max_row + 1):
            celda = ws.cell(row=row, column=col_password)
            celda.font = fuente_mono
            celda.number_format = "@"
            celda.alignment = Alignment(horizontal="left")

        for col in range(1, len(titulos) + 1):
            largo = max(
                [len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1)]
            )
            ws.column_dimensions[get_column_letter(col)].width = min(max(largo + 3, 12), 42)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        wb.save(ruta)
        return ruta.resolve()
