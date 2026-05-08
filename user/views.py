import base64
import os
import secrets
from datetime import timedelta
from email.message import EmailMessage
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import requests
from django.contrib.auth import authenticate
from django.db import IntegrityError
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST

from .jwt_utils import create_jwt, get_user_from_request
from django.contrib.auth.models import Group

from .models import AREA_CHOICES, TIPO_USUARIO_CHOICES, PasswordResetCode, Usuario


def _is_admin_user(user):
    if user is None:
        return False
    if getattr(user, "is_superuser", False) or getattr(user, "is_staff", False):
        return True
    groups = set(user.groups.values_list("name", flat=True))
    groups = {g.lower() for g in groups}
    return "admin" in groups


RESET_CODE_TTL_MINUTES = 10


def _generate_reset_code():
    return f"{secrets.randbelow(1_000_000):06d}"


def _get_gmail_access_token():
    client_id = os.environ.get("GMAIL_CLIENT_ID")
    client_secret = os.environ.get("GMAIL_CLIENT_SECRET")
    refresh_token = os.environ.get("GMAIL_REFRESH_TOKEN")
    if not client_id or not client_secret or not refresh_token:
        return None, "Faltan credenciales de Gmail en el .env."

    try:
        response = requests.post(
            "https://oauth2.googleapis.com/token",
            data={
                "client_id": client_id,
                "client_secret": client_secret,
                "refresh_token": refresh_token,
                "grant_type": "refresh_token",
            },
            timeout=10,
        )
    except requests.RequestException:
        return None, "No se pudo conectar con Gmail."

    if response.status_code != 200:
        return None, "No se pudo autenticar con Gmail."

    access_token = response.json().get("access_token")
    if not access_token:
        return None, "No se pudo obtener el token de Gmail."
    return access_token, None


def _send_reset_code_email(to_email, code):
    from_email = os.environ.get("GMAIL_FROM")
    if not from_email:
        return False, "Falta GMAIL_FROM en el .env."

    access_token, error = _get_gmail_access_token()
    if error:
        return False, error

    message = EmailMessage()
    message["To"] = to_email
    message["From"] = from_email
    message["Subject"] = "Codigo de recuperacion de contrasena"
    message.set_content(
        "Tu codigo de recuperacion es: "
        f"{code}\nEste codigo vence en {RESET_CODE_TTL_MINUTES} minutos."
    )

    raw = base64.urlsafe_b64encode(message.as_bytes()).decode("utf-8")
    try:
        response = requests.post(
            "https://gmail.googleapis.com/gmail/v1/users/me/messages/send",
            headers={"Authorization": f"Bearer {access_token}"},
            json={"raw": raw},
            timeout=10,
        )
    except requests.RequestException:
        return False, "No se pudo enviar el correo."

    if response.status_code not in (200, 202):
        return False, "No se pudo enviar el correo."
    return True, None


def crear_usuario(request):
    user = get_user_from_request(request)
    if user is None:
        return redirect("login")
    if not _is_admin_user(user):
        return redirect("home_autenticado")

    context = {
        "grupos": Group.objects.order_by("name"),
        "tipo_usuario_choices": TIPO_USUARIO_CHOICES,
        "area_choices": AREA_CHOICES,
    }
    if request.method == "POST":
        email = request.POST.get("email", "").strip()
        username = request.POST.get("username", "").strip()
        nombre = request.POST.get("nombre", "").strip()
        apellido = request.POST.get("apellido", "").strip()
        edad_raw = request.POST.get("edad", "").strip()
        telefono = request.POST.get("telefono", "").strip()
        tipo_usuario = request.POST.get("tipo_usuario", "").strip() or None
        area = request.POST.get("area", "").strip() or None
        password = request.POST.get("password", "")
        grupos_ids = request.POST.getlist("grupos")
        valid_tipo_usuario = {key for key, _ in TIPO_USUARIO_CHOICES}
        valid_area = {key for key, _ in AREA_CHOICES}

        if not email or not username or not nombre or not password:
            context["error"] = "Email, usuario, nombre y contraseña son obligatorios."
        else:
            edad_val = None
            if edad_raw:
                if edad_raw.isdigit():
                    edad_val = int(edad_raw)
                else:
                    context["error"] = "La edad debe ser un numero."

            if tipo_usuario and tipo_usuario not in valid_tipo_usuario:
                context["error"] = "El tipo de usuario no es valido."
            if area and area not in valid_area:
                context["error"] = "El area no es valida."

            if "error" not in context:
                try:
                    new_user = Usuario.objects.create_user(
                        email=email, username=username, password=password
                    )
                    new_user.nombre = nombre
                    new_user.apellido = apellido
                    new_user.edad = edad_val
                    new_user.telefono = telefono or None
                    new_user.tipo_usuario = tipo_usuario
                    new_user.area = area
                    new_user.save()
                    if grupos_ids:
                        grupos_validos = Group.objects.filter(id__in=grupos_ids)
                        new_user.groups.set(grupos_validos)
                    context["success"] = "Usuario creado correctamente."
                except IntegrityError:
                    context["error"] = "El email o usuario ya existe."

    return render(request, "crear_usuarios.html", context)


def login_view(request):
    if request.method == "POST":
        email = request.POST.get("email", "").strip()
        password = request.POST.get("password", "")
        user = authenticate(request, email=email, password=password)
        if user is None:
            return render(
                request,
                "login.html",
                {"error": "Credenciales incorrectas."},
            )
        token = create_jwt(user)
        response = redirect("home_autenticado")
        response.set_cookie(
            "jwt",
            token,
            httponly=True,
            samesite="Lax",
            secure=request.is_secure(),
            max_age=None,
        )
        return response
    return render(request, "login.html")


def recuperar_password(request):
    context = {}
    if request.method == "POST":
        action = request.POST.get("action", "").strip()
        email = request.POST.get("email", "").strip().lower()
        if email:
            context["email_value"] = email

        if action == "send_code":
            if not email:
                context["error"] = "El correo es obligatorio."
            else:
                user = Usuario.objects.filter(email__iexact=email).first()
                if user:
                    PasswordResetCode.objects.filter(
                        user=user, used_at__isnull=True, expires_at__gt=timezone.now()
                    ).update(used_at=timezone.now())

                    code = _generate_reset_code()
                    expires_at = timezone.now() + timedelta(minutes=RESET_CODE_TTL_MINUTES)
                    PasswordResetCode.objects.create(
                        user=user,
                        code=code,
                        expires_at=expires_at,
                    )
                    sent, error = _send_reset_code_email(email, code)
                    if not sent:
                        PasswordResetCode.objects.filter(
                            user=user, code=code, used_at__isnull=True
                        ).update(used_at=timezone.now())
                        context["error"] = error
                    else:
                        context["success"] = (
                            "Si el correo existe, se envio un codigo de recuperacion."
                        )
                else:
                    context["success"] = (
                        "Si el correo existe, se envio un codigo de recuperacion."
                    )

        elif action == "reset_password":
            code = request.POST.get("code", "").strip()
            new_password = request.POST.get("new_password", "")
            confirm = request.POST.get("confirm_password", "")
            context["code_value"] = code

            if not email:
                context["error_reset"] = "El correo es obligatorio."
            elif not code:
                context["error_reset"] = "El codigo es obligatorio."
            elif not code.isdigit() or len(code) != 6:
                context["error_reset"] = "El codigo debe tener 6 digitos."
            elif not new_password:
                context["error_reset"] = "La nueva contrasena es obligatoria."
            elif new_password != confirm:
                context["error_reset"] = "Las contrasenas no coinciden."
            else:
                user = Usuario.objects.filter(email__iexact=email).first()
                if not user:
                    context["error_reset"] = "Codigo o correo invalidos."
                else:
                    now = timezone.now()
                    reset = (
                        PasswordResetCode.objects.filter(
                            user=user,
                            code=code,
                            used_at__isnull=True,
                            expires_at__gt=now,
                        )
                        .order_by("-created_at")
                        .first()
                    )
                    if not reset:
                        context["error_reset"] = "Codigo invalido o vencido."
                    else:
                        user.set_password(new_password)
                        user.save()
                        reset.used_at = now
                        reset.save()
                        context["success_reset"] = "Contrasena actualizada."

    return render(request, "recuperar_password.html", context)


@require_POST
def logout_view(request):
    response = redirect("home")
    response.delete_cookie("jwt")
    return response


def settings_user(request):
    user = get_user_from_request(request)
    if user is None:
        return redirect("login")

    context = {"user_obj": user}

    if request.method == "POST":
        action = request.POST.get("action", "")

        if action == "update_info":
            user.username = request.POST.get("username", "").strip()
            user.nombre = request.POST.get("nombre", "").strip()
            user.apellido = request.POST.get("apellido", "").strip()
            edad_raw = request.POST.get("edad", "").strip()
            if edad_raw:
                if edad_raw.isdigit():
                    user.edad = int(edad_raw)
                else:
                    context["error_info"] = "La edad debe ser un numero."
            else:
                user.edad = None
            user.telefono = request.POST.get("telefono", "").strip() or None
            if "error_info" not in context:
                try:
                    user.save()
                    context["success_info"] = "Informacion actualizada."
                except IntegrityError:
                    context["error_info"] = "El usuario ya existe."

        elif action == "change_password":
            current = request.POST.get("current_password", "")
            new = request.POST.get("new_password", "")
            confirm = request.POST.get("confirm_password", "")

            if not user.check_password(current):
                context["error_password"] = "La contraseña actual no es correcta."
            elif not new:
                context["error_password"] = "La nueva contraseña es obligatoria."
            elif new != confirm:
                context["error_password"] = "Las contraseñas no coinciden."
            else:
                user.set_password(new)
                user.save()
                context["success_password"] = "Contraseña actualizada."

    return render(request, "settingsuser.html", context)


def home_user(request):
    user = get_user_from_request(request)
    if user is None:
        return redirect("login")
    return render(request, "home_user.html")


def listado_usuarios(request):
    user = get_user_from_request(request)
    if user is None:
        return redirect("login")
    if not _is_admin_user(user):
        return redirect("home_autenticado")

    context = {
        "usuarios": Usuario.objects.order_by("email"),
        "grupos": Group.objects.order_by("name"),
        "tipo_usuario_choices": TIPO_USUARIO_CHOICES,
        "area_choices": AREA_CHOICES,
    }

    if request.method == "POST":
        action = request.POST.get("action", "").strip()

        if action == "create_group":
            group_name = request.POST.get("group_name", "").strip()
            if not group_name:
                context["error"] = "El nombre del grupo es obligatorio."
            else:
                group_obj, created = Group.objects.get_or_create(name=group_name)
                if created:
                    context["success"] = "Grupo creado correctamente."
                else:
                    context["error"] = "El grupo ya existe."
        elif action == "create":
            email = request.POST.get("email", "").strip()
            username = request.POST.get("username", "").strip()
            nombre = request.POST.get("nombre", "").strip()
            apellido = request.POST.get("apellido", "").strip()
            edad_raw = request.POST.get("edad", "").strip()
            telefono = request.POST.get("telefono", "").strip()
            tipo_usuario = request.POST.get("tipo_usuario", "").strip() or None
            area = request.POST.get("area", "").strip() or None
            password = request.POST.get("password", "")
            grupos_ids = request.POST.getlist("grupos")
            valid_tipo_usuario = {key for key, _ in TIPO_USUARIO_CHOICES}
            valid_area = {key for key, _ in AREA_CHOICES}

            if not email or not username or not nombre or not password:
                context["error"] = "Email, usuario, nombre y contraseña son obligatorios."
            else:
                edad_val = None
                if edad_raw:
                    if edad_raw.isdigit():
                        edad_val = int(edad_raw)
                    else:
                        context["error"] = "La edad debe ser un numero."
                if tipo_usuario and tipo_usuario not in valid_tipo_usuario:
                    context["error"] = "El tipo de usuario no es valido."
                if area and area not in valid_area:
                    context["error"] = "El area no es valida."

            if "error" not in context:
                try:
                    new_user = Usuario.objects.create_user(
                        email=email, username=username, password=password
                    )
                    new_user.nombre = nombre
                    new_user.apellido = apellido
                    new_user.edad = edad_val
                    new_user.telefono = telefono or None
                    new_user.tipo_usuario = tipo_usuario
                    new_user.area = area
                    new_user.save()
                    if grupos_ids:
                        grupos_validos = Group.objects.filter(id__in=grupos_ids)
                        new_user.groups.set(grupos_validos)
                    context["success"] = "Usuario creado correctamente."
                except IntegrityError:
                    context["error"] = "El email o usuario ya existe."
        else:
            user_id = request.POST.get("user_id", "").strip()
            if not user_id:
                context["error"] = "Usuario no valido."
            else:
                try:
                    target = Usuario.objects.get(pk=user_id)
                except Usuario.DoesNotExist:
                    context["error"] = "Usuario no encontrado."
                else:
                    if action == "delete":
                        if target.id == user.id:
                            context["error"] = "No puedes eliminar tu propio usuario."
                        else:
                            target.delete()
                            context["success"] = "Usuario eliminado."
                        return render(request, "listado_usuarios.html", context)

                    email = request.POST.get("email", "").strip()
                    username = request.POST.get("username", "").strip()
                    nombre = request.POST.get("nombre", "").strip()
                    apellido = request.POST.get("apellido", "").strip()
                    edad_raw = request.POST.get("edad", "").strip()
                    telefono = request.POST.get("telefono", "").strip()
                    tipo_usuario = request.POST.get("tipo_usuario", "").strip() or None
                    area = request.POST.get("area", "").strip() or None
                    grupos_ids = request.POST.getlist("grupos")
                    valid_tipo_usuario = {key for key, _ in TIPO_USUARIO_CHOICES}
                    valid_area = {key for key, _ in AREA_CHOICES}

                    if not email or not username or not nombre:
                        context["error"] = "Email, usuario y nombre son obligatorios."
                    else:
                        edad_val = None
                        if edad_raw:
                            if edad_raw.isdigit():
                                edad_val = int(edad_raw)
                            else:
                                context["error"] = "La edad debe ser un numero."
                        if tipo_usuario and tipo_usuario not in valid_tipo_usuario:
                            context["error"] = "El tipo de usuario no es valido."
                        if area and area not in valid_area:
                            context["error"] = "El area no es valida."

                    if "error" not in context:
                        try:
                            target.email = email
                            target.username = username
                            target.nombre = nombre
                            target.apellido = apellido
                            target.edad = edad_val
                            target.telefono = telefono or None
                            target.tipo_usuario = tipo_usuario
                            target.area = area
                            target.save()

                            if grupos_ids:
                                grupos_validos = Group.objects.filter(id__in=grupos_ids)
                                target.groups.set(grupos_validos)
                            else:
                                target.groups.clear()

                            context["success"] = "Usuario actualizado."
                        except IntegrityError:
                            context["error"] = "El email o usuario ya existe."

        context["usuarios"] = Usuario.objects.order_by("email")
        context["grupos"] = Group.objects.order_by("name")

    return render(request, "listado_usuarios.html", context)


IMPORT_COLUMNS = [
    ("email", "Email (obligatorio)"),
    ("username", "Usuario (obligatorio)"),
    ("nombre", "Nombre (obligatorio)"),
    ("apellido", "Apellido"),
    ("edad", "Edad"),
    ("telefono", "Telefono"),
    ("tipo_usuario", "Tipo usuario"),
    ("area", "Area"),
    ("password", "Contrasena (obligatorio)"),
    ("grupos", "Grupos (separados por coma)"),
]


def descargar_plantilla_usuarios(request):
    user = get_user_from_request(request)
    if user is None:
        return redirect("login")
    if not _is_admin_user(user):
        return redirect("home_autenticado")

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Usuarios"

    header_font = Font(bold=True, color="FFFFFF", name="Calibri")
    header_fill = PatternFill(start_color="5934BC", end_color="5934BC", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    for idx, (_key, header) in enumerate(IMPORT_COLUMNS, start=1):
        cell = sheet.cell(row=1, column=idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        sheet.column_dimensions[get_column_letter(idx)].width = max(20, len(header) + 4)

    ejemplo = [
        "ejemplo@empresa.com",
        "ejemplo.user",
        "Juan",
        "Perez",
        30,
        "3001234567",
        "colaborador",
        "trade",
        "Cambiar123*",
        "Admin,Operaciones",
    ]
    for idx, valor in enumerate(ejemplo, start=1):
        sheet.cell(row=2, column=idx, value=valor)

    tipos_validos = ", ".join(key for key, _ in TIPO_USUARIO_CHOICES)
    areas_validas = ", ".join(key for key, _ in AREA_CHOICES)

    instrucciones = workbook.create_sheet(title="Instrucciones")
    instrucciones.column_dimensions["A"].width = 30
    instrucciones.column_dimensions["B"].width = 90
    titulo = instrucciones.cell(row=1, column=1, value="Instrucciones de importacion")
    titulo.font = Font(bold=True, size=14)
    instrucciones.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

    filas_info = [
        ("email", "Obligatorio. Debe ser unico."),
        ("username", "Obligatorio. Debe ser unico."),
        ("nombre", "Obligatorio."),
        ("apellido", "Opcional."),
        ("edad", "Opcional. Solo numero entero."),
        ("telefono", "Opcional."),
        ("tipo_usuario", f"Opcional. Valores validos: {tipos_validos}"),
        ("area", f"Opcional. Valores validos: {areas_validas}"),
        ("password", "Obligatorio. Contrasena inicial del usuario."),
        ("grupos", "Opcional. Nombres de grupos existentes separados por coma."),
    ]
    for idx, (campo, descripcion) in enumerate(filas_info, start=3):
        celda_a = instrucciones.cell(row=idx, column=1, value=campo)
        celda_a.font = Font(bold=True)
        instrucciones.cell(row=idx, column=2, value=descripcion)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        'attachment; filename="plantilla_importacion_usuarios.xlsx"'
    )
    return response


@require_POST
def importar_usuarios(request):
    user = get_user_from_request(request)
    if user is None:
        return redirect("login")
    if not _is_admin_user(user):
        return redirect("home_autenticado")

    archivo = request.FILES.get("archivo")
    context = {
        "usuarios": Usuario.objects.order_by("email"),
        "grupos": Group.objects.order_by("name"),
        "tipo_usuario_choices": TIPO_USUARIO_CHOICES,
        "area_choices": AREA_CHOICES,
    }

    if not archivo:
        context["error"] = "Debes seleccionar un archivo Excel."
        return render(request, "listado_usuarios.html", context)

    nombre_archivo = (archivo.name or "").lower()
    if not nombre_archivo.endswith(".xlsx"):
        context["error"] = "El archivo debe tener extension .xlsx."
        return render(request, "listado_usuarios.html", context)

    try:
        workbook = openpyxl.load_workbook(archivo, data_only=True)
    except Exception:
        context["error"] = "No se pudo leer el archivo Excel."
        return render(request, "listado_usuarios.html", context)

    sheet = workbook.active
    if sheet.max_row < 2:
        context["error"] = "El archivo no tiene filas con datos."
        return render(request, "listado_usuarios.html", context)

    valid_tipo_usuario = {key for key, _ in TIPO_USUARIO_CHOICES}
    valid_area = {key for key, _ in AREA_CHOICES}
    grupos_por_nombre = {g.name.lower(): g for g in Group.objects.all()}

    creados = 0
    errores = []

    column_keys = [key for key, _ in IMPORT_COLUMNS]

    for fila_idx, fila in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        if not any(celda not in (None, "") for celda in fila):
            continue

        datos = {}
        for idx, key in enumerate(column_keys):
            valor = fila[idx] if idx < len(fila) else None
            if isinstance(valor, str):
                valor = valor.strip()
            datos[key] = valor

        email = (datos.get("email") or "").strip() if datos.get("email") else ""
        username = (datos.get("username") or "").strip() if datos.get("username") else ""
        nombre = (datos.get("nombre") or "").strip() if datos.get("nombre") else ""
        password = datos.get("password")
        if isinstance(password, (int, float)):
            password = str(password)
        password = (password or "").strip() if password else ""

        if not email or not username or not nombre or not password:
            errores.append(
                f"Fila {fila_idx}: email, usuario, nombre y contrasena son obligatorios."
            )
            continue

        edad_val = None
        edad_raw = datos.get("edad")
        if edad_raw not in (None, ""):
            try:
                edad_val = int(edad_raw)
                if edad_val < 0:
                    raise ValueError
            except (TypeError, ValueError):
                errores.append(f"Fila {fila_idx}: la edad debe ser un numero entero.")
                continue

        tipo_usuario = datos.get("tipo_usuario")
        if tipo_usuario:
            tipo_usuario = str(tipo_usuario).strip().lower() or None
            if tipo_usuario and tipo_usuario not in valid_tipo_usuario:
                errores.append(
                    f"Fila {fila_idx}: tipo_usuario '{tipo_usuario}' no es valido."
                )
                continue
        else:
            tipo_usuario = None

        area = datos.get("area")
        if area:
            area = str(area).strip().lower() or None
            if area and area not in valid_area:
                errores.append(f"Fila {fila_idx}: area '{area}' no es valida.")
                continue
        else:
            area = None

        telefono = datos.get("telefono")
        if telefono not in (None, ""):
            telefono = str(telefono).strip() or None
        else:
            telefono = None

        apellido = datos.get("apellido")
        if apellido not in (None, ""):
            apellido = str(apellido).strip()
        else:
            apellido = ""

        grupos_raw = datos.get("grupos")
        grupos_objs = []
        grupos_invalidos = []
        if grupos_raw:
            nombres = [g.strip() for g in str(grupos_raw).split(",") if g.strip()]
            for nombre_grupo in nombres:
                grupo = grupos_por_nombre.get(nombre_grupo.lower())
                if grupo is None:
                    grupos_invalidos.append(nombre_grupo)
                else:
                    grupos_objs.append(grupo)
        if grupos_invalidos:
            errores.append(
                f"Fila {fila_idx}: grupos no encontrados: {', '.join(grupos_invalidos)}."
            )
            continue

        try:
            nuevo = Usuario.objects.create_user(
                email=email, username=username, password=password
            )
            nuevo.nombre = nombre
            nuevo.apellido = apellido
            nuevo.edad = edad_val
            nuevo.telefono = telefono
            nuevo.tipo_usuario = tipo_usuario
            nuevo.area = area
            nuevo.save()
            if grupos_objs:
                nuevo.groups.set(grupos_objs)
            creados += 1
        except IntegrityError:
            errores.append(
                f"Fila {fila_idx}: el email '{email}' o usuario '{username}' ya existe."
            )
        except Exception:
            errores.append(f"Fila {fila_idx}: no se pudo crear el usuario.")

    mensajes = []
    if creados:
        mensajes.append(f"Se importaron {creados} usuario(s) correctamente.")
    if errores:
        context["error"] = " | ".join([*mensajes, *errores])
    else:
        context["success"] = (
            mensajes[0] if mensajes else "No se importaron filas (archivo vacio)."
        )

    context["usuarios"] = Usuario.objects.order_by("email")
    context["grupos"] = Group.objects.order_by("name")
    return render(request, "listado_usuarios.html", context)
